import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook
import pickle
import os
import pulp
import numpy as np
import math
import hashlib
from collections import defaultdict
import json
import psycopg2

# 页面配置必须放在最前面
st.set_page_config(
    page_title="电价数据查询系统",
    page_icon="⚡",
    layout="wide"
)

st.title("⚡ 站点电价数据查询系统")

# 全局表格样式：数据与标题列居中对齐
st.markdown("""
<style>
    .stDataFrame th, .stDataFrame td {
        text-align: center !important;
    }
</style>
""", unsafe_allow_html=True)

# 定义脚本目录和电价数据目录
SCRIPT_DIR = Path(__file__).resolve().parent
PRICE_DATA_DIR = SCRIPT_DIR / "电价数据"
CACHE_DIR = SCRIPT_DIR / ".cache"
CACHE_DIR.mkdir(exist_ok=True)
RANKING_CACHE_FILE = CACHE_DIR / ".price_spread_rank_cache.pkl"
PROVINCIAL_AVG_CACHE_FILE = CACHE_DIR / ".provincial_avg_cache.pkl"
RANKING_CACHE_VERSION = 1
PROVINCIAL_AVG_CACHE_VERSION = 2
FACTORY_GROUP_COLUMN = "厂站类型"
FACTORY_STATION_LABEL = "电厂"
NON_FACTORY_STATION_LABEL = "电站"
GUANGDONG_SHARP_PEAK_MONTHS = {7, 8, 9}

# TimescaleDB 数据库连接配置
DB_CONFIG = {
    'dbname': 'postgres',
    'user': 'kehuitao',
    'password': '123456',
    'host': '192.168.100.8',
    'port': '5432'
}

# 储能优化参数配置
STORAGE_CONFIG = {
    'P': 250000,  # 储能逆变器功率，单位 kW
    'battery_capacity': 500000,  # 电池容量上限，单位 kWh
    'initial_soc': 0,  # 初始电量，单位 kWh
    'efficiency': 0.85,  # 放电效率
    'dt': 0.25,  # 时间间隔，小时
    'num': 96  # 每天时段数
}

# 使用单选切换视图，避免隐藏页面也执行耗时计算
view_mode = st.radio(
    "功能切换",
    options=["📊 电价数据查询", "📈 电价差排名", "🔋 储能配储优化"],
    horizontal=True,
    label_visibility="collapsed"
)

# 加载电站名.xlsx文件
@st.cache_data
def load_station_info():
    """加载电站名.xlsx文件，获取站点与母线的对应关系"""
    try:
        # 尝试多个可能的位置
        possible_paths = [
            SCRIPT_DIR / "电站名.xlsx",
            SCRIPT_DIR / "电价名.xlsx",
            PRICE_DATA_DIR / "电站名.xlsx",
            PRICE_DATA_DIR / "电价名.xlsx",
            SCRIPT_DIR.parent / "电站名.xlsx",
            Path.home() / "Desktop" / "电站名.xlsx",
            Path.home() / "Documents" / "电站名.xlsx",
        ]

        for path in possible_paths:
            if path.exists():
                df = pd.read_excel(path)
                return df

        return None
    except Exception:
        return None


# 数据库查询函数
def get_db_connection():
    """获取数据库连接"""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        return conn
    except Exception as e:
        st.error(f"数据库连接失败: {e}")
        return None


def check_db_available():
    """检测数据库是否可用，返回(True, None)或(False, 错误信息)"""
    try:
        conn = psycopg2.connect(**DB_CONFIG, connect_timeout=5)
        conn.close()
        return True, None
    except Exception as e:
        return False, str(e)


@st.cache_data(show_spinner=False)
def get_db_station_list():
    """从数据库获取所有站点列表"""
    conn = get_db_connection()
    if conn is None:
        return []
    
    try:
        query = """
            SELECT DISTINCT station
            FROM real_time_electricity_price
            ORDER BY station
        """
        df = pd.read_sql_query(query, conn)
        return df['station'].tolist()
    except Exception as e:
        st.error(f"查询站点列表失败: {e}")
        return []
    finally:
        conn.close()


@st.cache_data(show_spinner=False)
def get_db_available_years():
    """从数据库获取可用的年份列表"""
    conn = get_db_connection()
    if conn is None:
        return []
    
    try:
        query = """
            SELECT DISTINCT EXTRACT(YEAR FROM time)::integer AS year
            FROM real_time_electricity_price
            ORDER BY year DESC
        """
        df = pd.read_sql_query(query, conn)
        return df['year'].tolist()
    except Exception as e:
        st.error(f"查询年份列表失败: {e}")
        return []
    finally:
        conn.close()


@st.cache_data(show_spinner=False)
def load_price_data_from_db(station_name, year):
    """从数据库加载指定站点和年份的电价数据
    
    Args:
        station_name: 站点名称
        year: 年份 (int)
    
    Returns:
        DataFrame: 包含日期列和96个时段电价列的宽格式数据
    """
    conn = get_db_connection()
    if conn is None:
        return None
    
    try:
        start_date = f"{year}-01-01"
        end_date = f"{year + 1}-01-01"
        
        query = """
            SELECT time, price
            FROM real_time_electricity_price
            WHERE station = %s
              AND time >= %s AND time < %s
            ORDER BY time
        """
        df = pd.read_sql_query(query, conn, params=(station_name, start_date, end_date))
        
        if df.empty:
            return None
        
        # 转换时间格式
        df['time'] = pd.to_datetime(df['time'])
        df['date'] = df['time'].dt.date
        df['time_str'] = df['time'].dt.strftime('%H:%M')
        
        # 透视表：将长格式转换为宽格式
        pivot_df = df.pivot_table(
            index='date',
            columns='time_str',
            values='price',
            aggfunc='first'
        )
        
        # 重置索引并重命名列
        pivot_df = pivot_df.reset_index()
        pivot_df.columns.name = None
        
        # 按时间排序列（00:00, 00:15, 00:30, ...）
        time_cols = [col for col in pivot_df.columns if col != 'date']
        time_cols_sorted = sorted(time_cols)
        pivot_df = pivot_df[['date'] + time_cols_sorted]
        
        # 将date列转换为datetime类型
        pivot_df['date'] = pd.to_datetime(pivot_df['date'])
        
        return pivot_df
    except Exception as e:
        st.error(f"从数据库加载数据失败: {e}")
        return None
    finally:
        conn.close()


@st.cache_data(show_spinner="正在计算全省平均电价...")
def compute_provincial_average_from_db(station_year_pairs):
    """从数据库计算全省所有站点各时段平均电价
    
    Args:
        station_year_pairs: [(station_name, year), ...] 站点和年份的元组列表
    
    Returns:
        (date_list, time_columns, avg_by_date, failed_stations)
    """
    # 先检测数据库连接
    db_available, db_error = check_db_available()
    if not db_available:
        st.error(f"数据库连接断开: {db_error}")
        return None, None, None, []
    
    all_daily_arrays = []
    failed_stations = []
    
    for station_name, year in station_year_pairs:
        df = load_price_data_from_db(station_name, year)
        if df is None:
            failed_stations.append(station_name)
            continue
        
        if len(df.columns) < 2:
            failed_stations.append(station_name)
            continue
        
        date_col = df.columns[0]
        price_cols = df.columns[1:]
        
        if len(price_cols) < 96:
            failed_stations.append(station_name)
            continue
        price_cols = price_cols[:96]
        
        for _, row in df.iterrows():
            date_val = row[date_col]
            try:
                date_str = pd.Timestamp(date_val).strftime('%Y-%m-%d')
            except Exception:
                date_str = str(date_val)[:10]
            prices = pd.to_numeric(row[price_cols], errors='coerce').values.astype(float)
            if len(prices) == 96 and not np.all(np.isnan(prices)):
                all_daily_arrays.append((date_str, prices))
    
    if not all_daily_arrays:
        return None, None, None, failed_stations
    
    # Group by date and compute mean
    date_grouped = {}
    for date_str, prices in all_daily_arrays:
        if date_str not in date_grouped:
            date_grouped[date_str] = []
        date_grouped[date_str].append(prices)
    
    date_list = sorted(date_grouped.keys())
    avg_by_date = np.array([np.mean(date_grouped[d], axis=0) for d in date_list])
    
    time_columns = [f"{h:02d}:{m:02d}" for h in range(24) for m in (0, 15, 30, 45)]
    
    return date_list, time_columns, avg_by_date, failed_stations


def _calculate_station_stats_from_db(station_name, year):
    """从数据库计算单个站点的电价差统计"""
    df = load_price_data_from_db(station_name, year)
    if df is None or len(df.columns) < 2:
        return None
    
    price_data = df.iloc[:, 1:]  # 除日期列外的所有列
    if price_data.empty:
        return None
    
    daily_max = price_data.max(axis=1)
    daily_min = price_data.min(axis=1)
    daily_spread = daily_max - daily_min
    all_prices = price_data.to_numpy().flatten()
    
    return {
        '日均电价差': round(daily_spread.mean(), 4),
        '全年最高电价差': round(daily_spread.max(), 4),
        '全年最低电价差': round(daily_spread.min(), 4),
        '全年平均电价': round(all_prices.mean(), 4),
        '数据天数': len(df)
    }


@st.cache_data(show_spinner=False)
def calculate_all_stations_price_spread_from_db(station_year_pairs):
    """从数据库计算所有站点的电价差统计数据
    
    Args:
        station_year_pairs: [(station_name, year), ...] 站点和年份的元组列表
    
    Returns:
        (stats_df, failed_stations, cache_summary)
    """
    if not station_year_pairs:
        return pd.DataFrame(), (), {"cached_count": 0, "recomputed_count": 0}
    
    station_stats = []
    failed_stations = []
    
    for station_name, year in station_year_pairs:
        try:
            stats = _calculate_station_stats_from_db(station_name, year)
            if stats is None:
                failed_stations.append((station_name, "无数据"))
                continue
            
            station_stats.append({
                '站点名称': station_name,
                **stats
            })
        except Exception as e:
            failed_stations.append((station_name, str(e)))
    
    if not station_stats:
        return pd.DataFrame(), tuple(failed_stations), {
            "cached_count": 0,
            "recomputed_count": len(station_year_pairs)
        }
    
    stats_df = pd.DataFrame(station_stats)
    stats_df = stats_df.sort_values('日均电价差', ascending=False).reset_index(drop=True)
    stats_df['排名'] = range(1, len(stats_df) + 1)
    stats_df = stats_df[['排名', '站点名称', '日均电价差', '全年最高电价差', '全年最低电价差', '全年平均电价', '数据天数']]
    
    return stats_df, tuple(failed_stations), {
        "cached_count": 0,
        "recomputed_count": len(station_year_pairs)
    }


def export_db_data_to_excel(station_name, year, progress_callback=None):
    """将数据库中的电价数据导出到本地Excel文件
    
    Args:
        station_name: 站点名称
        year: 年份
        progress_callback: 进度回调函数 callback(current, total, message)
    
    Returns:
        (success, message, file_path)
    """
    try:
        # 从数据库加载数据
        if progress_callback:
            progress_callback(0, 100, f"正在从数据库加载 {station_name} 的数据...")
        
        df = load_price_data_from_db(station_name, year)
        if df is None:
            return False, f"无法从数据库加载 {station_name} 的 {year} 年数据", None
        
        if progress_callback:
            progress_callback(30, 100, "数据加载完成，正在准备保存...")
        
        # 创建保存目录
        save_dir = PRICE_DATA_DIR / str(year)
        save_dir.mkdir(parents=True, exist_ok=True)
        
        # 生成文件名（只包含站点名）
        file_name = f"{station_name}.xlsx"
        file_path = save_dir / file_name
        
        if progress_callback:
            progress_callback(50, 100, f"正在保存到 {file_name}...")
        
        # 保存到Excel（覆盖已有文件）
        df.to_excel(file_path, index=False, engine='openpyxl')
        
        if progress_callback:
            progress_callback(100, 100, "保存完成！")
        
        return True, f"成功导出到 {file_path}", str(file_path)
    except Exception as e:
        return False, f"导出失败: {str(e)}", None


def export_all_db_data_to_excel(station_list, year, progress_placeholder=None):
    """批量将数据库中的电价数据导出到本地Excel文件
    
    Args:
        station_list: 站点名称列表
        year: 年份
        progress_placeholder: Streamlit placeholder用于显示进度
    
    Returns:
        (success_count, fail_count, failed_stations)
    """
    success_count = 0
    fail_count = 0
    failed_stations = []
    total = len(station_list)
    
    # 创建进度条
    if progress_placeholder:
        progress_bar = progress_placeholder.progress(0)
        status_text = progress_placeholder.empty()
    
    for i, station_name in enumerate(station_list):
        # 更新进度
        progress = int((i / total) * 100)
        if progress_placeholder:
            progress_bar.progress(progress)
            status_text.text(f"正在处理 ({i+1}/{total}): {station_name}")
        
        # 导出单个站点
        success, message, _ = export_db_data_to_excel(station_name, year)
        if success:
            success_count += 1
        else:
            fail_count += 1
            failed_stations.append((station_name, message))
    
    # 完成
    if progress_placeholder:
        progress_bar.progress(100)
        status_text.text(f"导出完成！成功: {success_count}, 失败: {fail_count}")
    
    return success_count, fail_count, failed_stations


def sort_price_data_dirs(directories):
    """按年份倒序排列目录，非纯数字目录排在后面。"""
    def sort_key(path):
        folder_name = path.name.strip()
        if folder_name.isdigit():
            return (0, -int(folder_name))
        return (1, folder_name)

    return sorted(directories, key=sort_key)


def get_price_data_dir_options():
    """获取可用的电价数据年份目录；如果没有子目录，则回退到根目录。"""
    if not PRICE_DATA_DIR.exists():
        return []

    year_dirs = sort_price_data_dirs(
        [path for path in PRICE_DATA_DIR.iterdir() if path.is_dir()]
    )

    if year_dirs:
        return [(path.name, path) for path in year_dirs]

    return [("当前目录", PRICE_DATA_DIR)]


# 扫描所有电价数据文件
def scan_price_files(price_data_dir):
    """扫描指定目录下的所有电价Excel文件"""
    price_files = {}

    if not price_data_dir.exists():
        st.error(f"电价数据目录不存在: {price_data_dir}")
        return price_files

    for file_path in sorted(price_data_dir.iterdir()):
        if not file_path.is_file():
            continue
        if file_path.suffix.lower() not in {".xlsx", ".xls"}:
            continue
        if file_path.name.startswith("~$"):
            continue

        # 从文件名提取站点名称，兼容旧命名和“站点名.xlsx”两种格式
        file_name = file_path.stem
        if file_name in {"电站名", "电价名"}:
            continue

        station_name = (
            file_name
            .replace("电价数据（一年）", "")
            .replace("电价数据(一年)", "")
            .replace("电价数据", "")
            .strip()
        )

        if station_name:
            price_files[station_name] = file_path

    return price_files

# 加载电价数据
def get_file_cache_key(file_path):
    """生成文件缓存键，文件更新后自动失效缓存。"""
    file_stat = file_path.stat()
    return str(file_path), file_stat.st_mtime_ns, file_stat.st_size


@st.cache_data(show_spinner=False)
def load_price_data(file_path_str, modified_time_ns, file_size):
    """加载指定站点的电价数据"""
    try:
        df = pd.read_excel(file_path_str)
        return df
    except Exception as e:
        st.error(f"加载数据失败: {e}")
        return None


@st.cache_data(show_spinner="正在计算全省平均电价...")
def compute_provincial_average(file_keys):
    """计算全省所有站点各时段平均电价，返回(date_list, time_columns, avg_by_date)"""
    all_daily_arrays = []  # list of (date_string, 96-element array)
    date_col_name = None

    for file_path_str, modified_time_ns, file_size in file_keys:
        df = load_price_data(file_path_str, modified_time_ns, file_size)
        if df is None or len(df.columns) < 2:
            continue
        if date_col_name is None:
            date_col_name = df.columns[0]

        price_cols = df.columns[1:]
        if len(price_cols) < 96:
            continue
        price_cols = price_cols[:96]

        for _, row in df.iterrows():
            date_val = row[date_col_name]
            try:
                date_str = pd.Timestamp(date_val).strftime('%Y-%m-%d')
            except Exception:
                date_str = str(date_val)[:10]
            prices = pd.to_numeric(row[price_cols], errors='coerce').values.astype(float)
            if len(prices) == 96 and not np.all(np.isnan(prices)):
                all_daily_arrays.append((date_str, prices))

    if not all_daily_arrays:
        return None, None, None

    # Group by date and compute mean
    date_grouped = {}
    for date_str, prices in all_daily_arrays:
        if date_str not in date_grouped:
            date_grouped[date_str] = []
        date_grouped[date_str].append(prices)

    date_list = sorted(date_grouped.keys())
    avg_by_date = np.array([np.mean(date_grouped[d], axis=0) for d in date_list])

    time_columns = [f"{h:02d}:{m:02d}" for h in range(24) for m in (0, 15, 30, 45)]

    return date_list, time_columns, avg_by_date

def build_price_file_index(price_files_dict):
    """构造稳定的文件索引，便于缓存排名结果。"""
    file_index = []
    for station_name in sorted(price_files_dict):
        file_path = price_files_dict[station_name]
        file_path_str, modified_time_ns, file_size = get_file_cache_key(file_path)
        file_index.append((station_name, file_path_str, modified_time_ns, file_size))
    return tuple(file_index)


def load_provincial_avg_cache():
    """读取全省平均电价的磁盘缓存。"""
    if not PROVINCIAL_AVG_CACHE_FILE.exists():
        return {}
    
    try:
        with open(PROVINCIAL_AVG_CACHE_FILE, "rb") as cache_file:
            cache_payload = pickle.load(cache_file)
        
        if cache_payload.get("version") != PROVINCIAL_AVG_CACHE_VERSION:
            return {}
        
        return cache_payload.get("entries", {})
    except Exception:
        return {}


def save_provincial_avg_cache(cache_data):
    """将全省平均电价缓存写入磁盘。"""
    cache_payload = {
        "version": PROVINCIAL_AVG_CACHE_VERSION,
        "entries": cache_data
    }
    temp_cache_file = PROVINCIAL_AVG_CACHE_FILE.with_suffix(".tmp")
    
    with open(temp_cache_file, "wb") as cache_file:
        pickle.dump(cache_payload, cache_file, protocol=pickle.HIGHEST_PROTOCOL)
    
    os.replace(temp_cache_file, PROVINCIAL_AVG_CACHE_FILE)


def _provincial_avg_cache_key(file_keys):
    """根据文件索引生成稳定的缓存键（MD5），用于区分不同年份/目录。"""
    key_parts = [f"{fp}|{mtime}|{size}" for fp, mtime, size in file_keys]
    raw = "\n".join(key_parts)
    return hashlib.md5(raw.encode("utf-8")).hexdigest()


def _is_provincial_avg_entry_valid(cached_entry, file_keys):
    """校验单条全省平均电价缓存条目是否仍然有效。"""
    if not isinstance(cached_entry, dict):
        return False

    cached_file_keys = cached_entry.get("file_keys", [])
    if len(cached_file_keys) != len(file_keys):
        return False

    for cached_key, current_key in zip(cached_file_keys, file_keys):
        if cached_key[1] != current_key[1] or cached_key[2] != current_key[2]:
            return False

    return True


def compute_provincial_average_with_cache(file_keys):
    """带磁盘缓存的全省平均电价计算，支持多年份缓存共存。"""
    cache_key = _provincial_avg_cache_key(file_keys)
    disk_cache = load_provincial_avg_cache()

    # 按哈希键查找对应条目
    cached_entry = disk_cache.get(cache_key)
    if cached_entry is not None and _is_provincial_avg_entry_valid(cached_entry, file_keys):
        date_list = cached_entry["date_list"]
        time_columns = cached_entry["time_columns"]
        avg_by_date = np.array(cached_entry["avg_by_date"])
        return date_list, time_columns, avg_by_date

    # 缓存未命中，重新计算
    date_list, time_columns, avg_by_date = compute_provincial_average(file_keys)

    if date_list is not None and time_columns is not None and avg_by_date is not None:
        # 将新条目合并到已有缓存中（保留其他年份的缓存）
        disk_cache[cache_key] = {
            "file_keys": file_keys,
            "date_list": date_list,
            "time_columns": time_columns,
            "avg_by_date": avg_by_date.tolist()
        }
        save_provincial_avg_cache(disk_cache)

    return date_list, time_columns, avg_by_date


def load_ranking_cache_entries():
    """读取磁盘缓存，支持跨 Streamlit 重启复用。"""
    if not RANKING_CACHE_FILE.exists():
        return {}

    try:
        with open(RANKING_CACHE_FILE, "rb") as cache_file:
            cache_payload = pickle.load(cache_file)

        if cache_payload.get("version") != RANKING_CACHE_VERSION:
            return {}

        cache_entries = cache_payload.get("entries", {})
        return cache_entries if isinstance(cache_entries, dict) else {}
    except Exception:
        return {}


def is_valid_ranking_cache_entry(cached_entry, modified_time_ns, file_size):
    """判断缓存条目是否仍可直接复用。"""
    return (
        cached_entry
        and cached_entry.get("modified_time_ns") == modified_time_ns
        and cached_entry.get("file_size") == file_size
        and isinstance(cached_entry.get("stats"), dict)
    )


def save_ranking_cache_entries(cache_entries):
    """将排名统计缓存写入磁盘。"""
    cache_payload = {
        "version": RANKING_CACHE_VERSION,
        "entries": cache_entries
    }
    temp_cache_file = RANKING_CACHE_FILE.with_suffix(".tmp")

    with open(temp_cache_file, "wb") as cache_file:
        pickle.dump(cache_payload, cache_file, protocol=pickle.HIGHEST_PROTOCOL)

    os.replace(temp_cache_file, RANKING_CACHE_FILE)


def merge_ranking_cache_entries(disk_cache_entries, current_cache_entries, price_file_index):
    """合并当前批次缓存，同时保留其他年份目录的缓存条目。"""
    merged_cache_entries = dict(disk_cache_entries)
    current_file_paths = {file_path_str for _, file_path_str, _, _ in price_file_index}
    current_parent_dirs = {str(Path(file_path_str).parent) for file_path_str in current_file_paths}

    if len(current_parent_dirs) == 1:
        current_parent_dir = next(iter(current_parent_dirs))
        merged_cache_entries = {
            cache_path: cache_entry
            for cache_path, cache_entry in merged_cache_entries.items()
            if str(Path(cache_path).parent) != current_parent_dir or cache_path in current_file_paths
        }

    merged_cache_entries.update(current_cache_entries)
    return merged_cache_entries


def _coerce_numeric(value):
    """将单元格值尽量转为浮点数。"""
    if value is None or value == "" or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


def _calculate_station_stats_with_openpyxl(file_path_str):
    """使用只读方式扫描 Excel，避免为排名统计创建整张 DataFrame。"""
    workbook = None
    try:
        workbook = load_workbook(file_path_str, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        next(rows, None)  # 跳过表头

        spread_sum = 0.0
        spread_max = None
        spread_min = None
        day_count = 0
        total_price_sum = 0.0
        total_price_count = 0

        for row in rows:
            row_min = None
            row_max = None
            row_sum = 0.0
            row_count = 0

            for cell_value in row[1:]:
                numeric_value = _coerce_numeric(cell_value)
                if numeric_value is None:
                    continue

                row_sum += numeric_value
                row_count += 1
                row_min = numeric_value if row_min is None else min(row_min, numeric_value)
                row_max = numeric_value if row_max is None else max(row_max, numeric_value)

            if row_count == 0:
                continue

            daily_spread = row_max - row_min
            spread_sum += daily_spread
            spread_max = daily_spread if spread_max is None else max(spread_max, daily_spread)
            spread_min = daily_spread if spread_min is None else min(spread_min, daily_spread)
            total_price_sum += row_sum
            total_price_count += row_count
            day_count += 1

        if day_count == 0 or total_price_count == 0:
            return None

        return {
            '日均电价差': round(spread_sum / day_count, 4),
            '全年最高电价差': round(spread_max, 4),
            '全年最低电价差': round(spread_min, 4),
            '全年平均电价': round(total_price_sum / total_price_count, 4),
            '数据天数': day_count
        }
    finally:
        if workbook is not None:
            workbook.close()


def _calculate_station_stats_with_pandas(file_path_str):
    """兼容 .xls 文件的统计逻辑。"""
    df = pd.read_excel(file_path_str)
    price_data = df.iloc[:, 1:]

    if price_data.empty:
        return None

    daily_max = price_data.max(axis=1)
    daily_min = price_data.min(axis=1)
    daily_spread = daily_max - daily_min
    all_prices = price_data.to_numpy().flatten()

    return {
        '日均电价差': round(daily_spread.mean(), 4),
        '全年最高电价差': round(daily_spread.max(), 4),
        '全年最低电价差': round(daily_spread.min(), 4),
        '全年平均电价': round(all_prices.mean(), 4),
        '数据天数': len(df)
    }


def calculate_station_price_stats(file_meta):
    """计算单个站点的统计值。"""
    station_name, file_path_str, modified_time_ns, file_size = file_meta

    try:
        file_suffix = Path(file_path_str).suffix.lower()
        if file_suffix == ".xlsx":
            stats = _calculate_station_stats_with_openpyxl(file_path_str)
        else:
            stats = _calculate_station_stats_with_pandas(file_path_str)

        if stats is None:
            return None

        return {
            '站点名称': station_name,
            **stats
        }
    except Exception as e:
        return {
            '站点名称': station_name,
            '__error__': str(e)
        }


# 计算所有站点的电价差统计
@st.cache_data(show_spinner=False)
def calculate_all_stations_price_spread(price_file_index):
    """并行计算所有站点的电价差统计数据，并复用磁盘缓存。"""
    if not price_file_index:
        return pd.DataFrame(), (), {"cached_count": 0, "recomputed_count": 0}

    disk_cache_entries = load_ranking_cache_entries()
    next_cache_entries = {}
    station_stats = []
    failed_stations = []
    pending_file_index = []
    cached_count = 0

    for file_meta in price_file_index:
        station_name, file_path_str, modified_time_ns, file_size = file_meta
        cached_entry = disk_cache_entries.get(file_path_str)

        if is_valid_ranking_cache_entry(cached_entry, modified_time_ns, file_size):
            cached_stats = {
                "站点名称": station_name,
                **cached_entry["stats"]
            }
            station_stats.append(cached_stats)
            next_cache_entries[file_path_str] = cached_entry
            cached_count += 1
            continue

        pending_file_index.append(file_meta)

    max_workers = min(len(pending_file_index), max(4, min(12, (os.cpu_count() or 4) + 2))) if pending_file_index else 0

    if pending_file_index:
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            results = list(executor.map(calculate_station_price_stats, pending_file_index))
    else:
        results = []

    for file_meta, result in zip(pending_file_index, results):
        station_name, file_path_str, modified_time_ns, file_size = file_meta
        if not result:
            continue
        if '__error__' in result:
            failed_stations.append((result['站点名称'], result['__error__']))
            continue

        station_stats.append(result)
        next_cache_entries[file_path_str] = {
            "modified_time_ns": modified_time_ns,
            "file_size": file_size,
            "stats": {
                "日均电价差": result["日均电价差"],
                "全年最高电价差": result["全年最高电价差"],
                "全年最低电价差": result["全年最低电价差"],
                "全年平均电价": result["全年平均电价"],
                "数据天数": result["数据天数"]
            }
        }

    if not station_stats:
        return pd.DataFrame(), tuple(failed_stations), {
            "cached_count": cached_count,
            "recomputed_count": len(pending_file_index)
        }

    merged_cache_entries = merge_ranking_cache_entries(
        disk_cache_entries,
        next_cache_entries,
        price_file_index,
    )

    if merged_cache_entries != disk_cache_entries:
        save_ranking_cache_entries(merged_cache_entries)

    stats_df = pd.DataFrame(station_stats)
    stats_df = stats_df.sort_values('日均电价差', ascending=False).reset_index(drop=True)
    stats_df['排名'] = range(1, len(stats_df) + 1)
    stats_df = stats_df[['排名', '站点名称', '日均电价差', '全年最高电价差', '全年最低电价差', '全年平均电价', '数据天数']]
    return stats_df, tuple(failed_stations), {
        "cached_count": cached_count,
        "recomputed_count": len(pending_file_index)
    }





def classify_factory_station_group(station_name):
    """根据站点名称判断是否属于电厂。"""
    station_name_str = str(station_name).strip()
    return FACTORY_STATION_LABEL if "厂" in station_name_str else NON_FACTORY_STATION_LABEL


def get_guangdong_period_type(date_value, time_str):
    """按广东省峰谷分时电价规则判定时段类型。"""
    try:
        hour_str, minute_str = str(time_str).split(':', 1)
        minutes = int(hour_str) * 60 + int(minute_str)
    except (TypeError, ValueError):
        return "未知"

    date_ts = pd.to_datetime(date_value, errors='coerce')
    # 广东尖峰还包括其他月份的高温天，但程序当前未接入气象数据，
    # 因此先按固定的 7-9 月整月识别尖峰时段。
    is_sharp_peak_month = pd.notna(date_ts) and date_ts.month in GUANGDONG_SHARP_PEAK_MONTHS

    if is_sharp_peak_month and (
        11 * 60 <= minutes < 12 * 60 or
        15 * 60 <= minutes < 17 * 60
    ):
        return "尖峰"

    if 10 * 60 <= minutes < 12 * 60 or 14 * 60 <= minutes < 19 * 60:
        return "高峰"

    if 0 <= minutes < 8 * 60:
        return "低谷"

    return "平段"


def build_station_group_mapping(station_info_df, group_column, station_names=None):
    """构造电站与分组字段的映射表。"""
    if group_column == FACTORY_GROUP_COLUMN:
        if station_names is None:
            return None

        station_names_list = (
            pd.Series(list(station_names))
            .dropna()
            .astype(str)
            .str.strip()
        )
        station_names_list = station_names_list[station_names_list != ""].drop_duplicates().tolist()

        if not station_names_list:
            return None

        return pd.DataFrame({
            '电站名': station_names_list,
            group_column: [classify_factory_station_group(name) for name in station_names_list]
        })

    if (
        station_info_df is None
        or '电站名' not in station_info_df.columns
        or group_column not in station_info_df.columns
    ):
        return None

    station_group_df = (
        station_info_df[['电站名', group_column]]
        .copy()
        .dropna(subset=['电站名'])
    )
    station_group_df['电站名'] = station_group_df['电站名'].astype(str).str.strip()
    station_group_df[group_column] = (
        station_group_df[group_column]
        .fillna('未分组')
        .astype(str)
        .str.strip()
        .replace('', '未分组')
    )
    station_group_df = station_group_df.drop_duplicates(subset=['电站名'], keep='first')
    return station_group_df


def get_available_group_values(station_info_df, group_column, station_names=None):
    """获取指定字段下可用的分组选项。"""
    station_group_df = build_station_group_mapping(station_info_df, group_column, station_names=station_names)
    if station_group_df is None:
        return []

    if group_column == FACTORY_GROUP_COLUMN:
        ordered_values = [FACTORY_STATION_LABEL, NON_FACTORY_STATION_LABEL]
        available_values = set(station_group_df[group_column].dropna().tolist())
        return [value for value in ordered_values if value in available_values]

    return sorted(station_group_df[group_column].dropna().unique().tolist())


def get_station_group_value(station_info_df, station_name, group_column):
    """获取单个站点对应的分组字段值。"""
    station_group_df = build_station_group_mapping(
        station_info_df,
        group_column,
        station_names=[station_name]
    )
    if station_group_df is None:
        return None

    station_name_str = str(station_name).strip()
    matched_df = station_group_df[station_group_df['电站名'] == station_name_str]
    if matched_df.empty:
        return None

    group_value = str(matched_df.iloc[0][group_column]).strip()
    if not group_value or group_value == '未分组':
        return None

    return group_value


def prepare_grouped_rankings(all_stations_stats, station_info_df, group_column):
    """为排名结果补充分组信息和组内排名。"""
    station_group_df = build_station_group_mapping(
        station_info_df,
        group_column,
        station_names=all_stations_stats['站点名称'].tolist()
    )
    if station_group_df is None:
        return None

    grouped_df = all_stations_stats.merge(
        station_group_df,
        left_on='站点名称',
        right_on='电站名',
        how='left'
    ).drop(columns=['电站名'])

    grouped_df[group_column] = grouped_df[group_column].fillna('未分组')
    grouped_df = grouped_df.sort_values(
        [group_column, '日均电价差', '站点名称'],
        ascending=[True, False, True]
    ).reset_index(drop=True)
    grouped_df['组内排名'] = grouped_df.groupby(group_column).cumcount() + 1
    grouped_df[f'{group_column}站点数'] = grouped_df.groupby(group_column)['站点名称'].transform('count')

    return grouped_df


# 储能优化函数
def optimize_single_day(price, day_index, start_soc, config):
    """优化单天的储能调度"""
    num = config['num']
    dt = config['dt']
    P = config['P']
    battery_capacity = config['battery_capacity']
    efficiency = config['efficiency']
    
    # 创建问题
    prob = pulp.LpProblem(f"BESS_Optimization_Day{day_index + 1}", pulp.LpMaximize)
    
    # 定义变量
    charge_power = pulp.LpVariable.dicts("charge", range(num), lowBound=0)
    discharge_power = pulp.LpVariable.dicts("discharge", range(num), lowBound=0)
    soc = pulp.LpVariable.dicts("soc", range(num + 1), lowBound=0, upBound=battery_capacity)
    
    # 二进制变量
    z_ch = pulp.LpVariable.dicts("is_charging", range(num), cat="Binary")
    z_dis = pulp.LpVariable.dicts("is_discharging", range(num), cat="Binary")
    y_ch = pulp.LpVariable.dicts("charge_start", range(num), cat="Binary")
    y_dis = pulp.LpVariable.dicts("discharge_start", range(num), cat="Binary")
    
    epsilon = 1e-3
    min_duration = 4
    min_on_power = 0.05 * P
    
    # 目标函数
    prob += pulp.lpSum([(discharge_power[i] - charge_power[i]) * price[i] * dt for i in range(num)])
    
    # 初始电量约束
    prob += soc[0] == start_soc
    
    # 电池状态转移方程
    for i in range(num):
        prob += soc[i + 1] == soc[i] + charge_power[i] * dt - discharge_power[i] * dt / efficiency
    
    # 功率-状态关联约束
    for i in range(num):
        prob += charge_power[i] <= P * z_ch[i]
        prob += discharge_power[i] <= P * z_dis[i]
        prob += charge_power[i] >= min_on_power * z_ch[i]
        prob += discharge_power[i] >= min_on_power * z_dis[i]
        prob += z_ch[i] + z_dis[i] <= 1
        
        if i == 0:
            prob += y_ch[i] == z_ch[i]
            prob += y_dis[i] == z_dis[i]
        else:
            prob += y_ch[i] >= z_ch[i] - z_ch[i-1]
            prob += y_ch[i] <= z_ch[i]
            prob += y_ch[i] <= 1 - z_ch[i-1]
            prob += y_dis[i] >= z_dis[i] - z_dis[i-1]
            prob += y_dis[i] <= z_dis[i]
            prob += y_dis[i] <= 1 - z_dis[i-1]
    
    # 最小连续时长约束
    for i in range(num):
        if i <= num - min_duration:
            prob += pulp.lpSum(z_ch[j] for j in range(i, i + min_duration)) >= min_duration * y_ch[i]
            prob += pulp.lpSum(z_dis[j] for j in range(i, i + min_duration)) >= min_duration * y_dis[i]
        else:
            prob += y_ch[i] == 0
            prob += y_dis[i] == 0
    
    # 累计能量平衡约束（允许电量留存）
    for k in range(1, num+1):
        prob += efficiency * pulp.lpSum([charge_power[i] * dt for i in range(k)]) + start_soc >= pulp.lpSum(
            [discharge_power[i] * dt for i in range(k)])
    
    # 连续充电能量约束
    M = int(math.ceil(battery_capacity / (P * dt))) + 1
    for s in range(0, 97 - M):
        end_idx = min(s + M - 1, 95)
        prob += pulp.lpSum([charge_power[j] * dt for j in range(s, end_idx + 1)]) <= battery_capacity
    
    # 求解
    solver = pulp.PULP_CBC_CMD(msg=False, timeLimit=120)
    result = prob.solve(solver)
    
    return prob, result


@st.cache_data(show_spinner="正在优化储能调度...")
def run_optimization_cached(file_path_or_none, P, battery_capacity, efficiency):
    """缓存化的多天储能优化，相同参数+相同文件只计算一次"""
    config = {
        'P': P, 'battery_capacity': battery_capacity,
        'initial_soc': 0, 'efficiency': efficiency,
        'dt': 0.25, 'num': 96
    }
    df = pd.read_excel(file_path_or_none)
    start_col = 1 if len(df.columns) > 96 and not pd.api.types.is_numeric_dtype(df.iloc[:, 0]) else 0
    all_prices = df.iloc[:, start_col:start_col + 96].values.astype(float)

    date_list = []
    if start_col == 1:
        for d in df.iloc[:, 0].values:
            if pd.isna(d):
                date_list.append('')
            elif isinstance(d, pd.Timestamp):
                date_list.append(d.strftime('%Y-%m-%d'))
            elif isinstance(d, str):
                ds = str(d).strip()
                date_list.append(ds.split('T')[0] if 'T' in ds else (ds.split(' ')[0] if ' ' in ds else ds[:10]))
            else:
                date_list.append(str(d)[:10])
    else:
        date_list = [f'第{i+1}天' for i in range(len(df))]

    num_days = len(all_prices)
    all_results = []
    all_summaries = []
    current_soc = 0.0

    for day_idx in range(num_days):
        price = all_prices[day_idx, :]
        day_start_soc = current_soc
        prob, result = optimize_single_day(price, day_idx, current_soc, config)

        if pulp.LpStatus[prob.status] == "Optimal":
            var_dict = {v.name: v for v in prob.variables()}
            charge_power_values = [float(pulp.value(var_dict[f"charge_{i}"])) for i in range(96)]
            discharge_power_values = [float(pulp.value(var_dict[f"discharge_{i}"])) for i in range(96)]
            soc_values = [float(pulp.value(var_dict[f"soc_{i}"])) for i in range(97)]
            current_soc = float(pulp.value(var_dict[f"soc_{96}"]))
            total_revenue = float(pulp.value(prob.objective))

            dt = 0.25
            for i in range(96):
                hour = i * dt
                time_str = f"{int(hour):02d}:{int((hour - int(hour)) * 60):02d}"
                if charge_power_values[i] > 1e-4:
                    period_type = "充电"
                elif discharge_power_values[i] > 1e-4:
                    period_type = "放电"
                else:
                    period_type = "空闲"
                all_results.append({
                    '日期': date_list[day_idx], '时间': time_str,
                    '电价_元/kWh': float(price[i]),
                    '充电功率_kW': charge_power_values[i],
                    '放电功率_kW': discharge_power_values[i],
                    '净功率_kW': discharge_power_values[i] - charge_power_values[i],
                    '电池电量_kWh': soc_values[i + 1],
                    '时段类型': period_type
                })

            total_charge = sum(charge_power_values[i] * dt for i in range(96))
            total_discharge = sum(discharge_power_values[i] * dt for i in range(96))
            all_summaries.append({
                '日期': date_list[day_idx],
                '日收益_元': total_revenue,
                '初始电量_kWh': day_start_soc,
                '最终电量_kWh': current_soc,
                '充电量_kWh': total_charge,
                '放电量_kWh': total_discharge
            })

    return date_list, all_results, all_summaries


@st.cache_data(show_spinner="正在优化储能调度...")
def run_optimization_cached_from_db(station_name, year, P, battery_capacity, efficiency):
    """从数据库加载数据并进行储能优化"""
    config = {
        'P': P, 'battery_capacity': battery_capacity,
        'initial_soc': 0, 'efficiency': efficiency,
        'dt': 0.25, 'num': 96
    }
    
    # 从数据库加载数据
    df = load_price_data_from_db(station_name, year)
    if df is None:
        raise ValueError(f"无法从数据库加载 {station_name} 的 {year} 年数据")
    
    start_col = 1 if len(df.columns) > 96 and not pd.api.types.is_numeric_dtype(df.iloc[:, 0]) else 0
    all_prices = df.iloc[:, start_col:start_col + 96].values.astype(float)

    date_list = []
    if start_col == 1:
        for d in df.iloc[:, 0].values:
            if pd.isna(d):
                date_list.append('')
            elif isinstance(d, pd.Timestamp):
                date_list.append(d.strftime('%Y-%m-%d'))
            elif isinstance(d, str):
                ds = str(d).strip()
                date_list.append(ds.split('T')[0] if 'T' in ds else (ds.split(' ')[0] if ' ' in ds else ds[:10]))
            else:
                date_list.append(str(d)[:10])
    else:
        date_list = [f'第{i+1}天' for i in range(len(df))]

    num_days = len(all_prices)
    all_results = []
    all_summaries = []
    current_soc = 0.0

    for day_idx in range(num_days):
        price = all_prices[day_idx, :]
        day_start_soc = current_soc
        prob, result = optimize_single_day(price, day_idx, current_soc, config)

        if pulp.LpStatus[prob.status] == "Optimal":
            var_dict = {v.name: v for v in prob.variables()}
            charge_power_values = [float(pulp.value(var_dict[f"charge_{i}"])) for i in range(96)]
            discharge_power_values = [float(pulp.value(var_dict[f"discharge_{i}"])) for i in range(96)]
            soc_values = [float(pulp.value(var_dict[f"soc_{i}"])) for i in range(97)]
            current_soc = float(pulp.value(var_dict[f"soc_{96}"]))
            total_revenue = float(pulp.value(prob.objective))

            dt = 0.25
            for i in range(96):
                hour = i * dt
                time_str = f"{int(hour):02d}:{int((hour - int(hour)) * 60):02d}"
                if charge_power_values[i] > 1e-4:
                    period_type = "充电"
                elif discharge_power_values[i] > 1e-4:
                    period_type = "放电"
                else:
                    period_type = "空闲"
                all_results.append({
                    '日期': date_list[day_idx], '时间': time_str,
                    '电价_元/kWh': float(price[i]),
                    '充电功率_kW': charge_power_values[i],
                    '放电功率_kW': discharge_power_values[i],
                    '净功率_kW': discharge_power_values[i] - charge_power_values[i],
                    '电池电量_kWh': soc_values[i + 1],
                    '时段类型': period_type
                })

            total_charge = sum(charge_power_values[i] * dt for i in range(96))
            total_discharge = sum(discharge_power_values[i] * dt for i in range(96))
            all_summaries.append({
                '日期': date_list[day_idx],
                '日收益_元': total_revenue,
                '初始电量_kWh': day_start_soc,
                '最终电量_kWh': current_soc,
                '充电量_kWh': total_charge,
                '放电量_kWh': total_discharge
            })

    return date_list, all_results, all_summaries


# 主程序
def main():
    # 加载电站信息
    station_info_df = load_station_info()

    # 数据源选择
    st.sidebar.header("📂 数据源配置")
    data_source = st.sidebar.radio(
        "选择数据源",
        options=["数据库", "Excel文件"],
        index=0,
        help="选择从数据库或Excel文件读取电价数据"
    )

    # 检测数据库可用性
    db_available, db_error = check_db_available() if data_source == "数据库" else (False, None)
    
    # 数据库不可用时自动切换到Excel模式
    if data_source == "数据库" and not db_available:
        st.warning(f"数据库连接失败: {db_error}")
        st.info("已自动切换到Excel文件模式，请确保电价数据目录中有数据文件。")
        data_source = "Excel文件"

    if data_source == "数据库":
        # 从数据库加载数据
        db_stations = get_db_station_list()
        if not db_stations:
            st.warning("数据库中未找到站点数据！")
            st.info("请检查数据库连接配置或确认数据库中已有电价数据。")
            return
        
        db_years = get_db_available_years()
        if not db_years:
            st.warning("数据库中未找到可用的年份数据！")
            return
        
        # 年份选择
        selected_year = st.sidebar.selectbox(
            "选择年份",
            options=db_years,
            index=0,
            help="选择要查询的年份"
        )
        
        # 构建price_files字典（兼容现有逻辑）
        # 使用站点名称作为键，None作为文件路径（因为从数据库加载）
        price_files = {station: None for station in db_stations}
        
        # 保存数据源信息供后续使用
        st.session_state['data_source'] = 'database'
        st.session_state['selected_year'] = selected_year
        
        # 数据同步功能
        st.sidebar.divider()
        st.sidebar.subheader("💾 数据同步")
        
        # 选择要同步的站点
        sync_mode = st.sidebar.radio(
            "同步范围",
            options=["全部站点", "选择站点"],
            horizontal=True,
            help="选择要同步到本地的站点范围"
        )
        
        if sync_mode == "选择站点":
            sync_stations = st.sidebar.multiselect(
                "选择站点",
                options=db_stations,
                default=[],
                help="选择要同步的站点"
            )
        else:
            sync_stations = db_stations
        
        # 同步按钮
        if st.sidebar.button("📤 同步到本地Excel", type="primary", disabled=len(sync_stations) == 0):
            st.sidebar.divider()
            st.sidebar.subheader("📊 同步进度")
            progress_placeholder = st.sidebar.empty()
            
            # 执行同步
            success_count, fail_count, failed_stations = export_all_db_data_to_excel(
                sync_stations, selected_year, progress_placeholder
            )
            
            # 显示结果
            if fail_count == 0:
                st.sidebar.success(f"✅ 同步完成！成功导出 {success_count} 个站点的数据")
            else:
                st.sidebar.warning(f"⚠️ 同步完成：成功 {success_count} 个，失败 {fail_count} 个")
                if failed_stations:
                    with st.sidebar.expander("查看失败详情"):
                        for station, error in failed_stations[:10]:
                            st.write(f"- {station}: {error}")
                        if len(failed_stations) > 10:
                            st.write(f"... 还有 {len(failed_stations) - 10} 个站点")
            
            # 同步后刷新文件列表
            st.sidebar.button("🔄 刷新文件列表", on_click=st.cache_data.clear)
        
    else:
        # 从Excel文件加载数据（原有逻辑）
        price_dir_options = get_price_data_dir_options()
        if not price_dir_options:
            st.warning("未找到电价数据目录！")
            st.info(f"请先创建电价数据目录: `{PRICE_DATA_DIR}`")
            return

        price_dir_labels = [label for label, _ in price_dir_options]
        default_price_dir_index = 0

        selected_price_dir_label = st.sidebar.selectbox(
            "电价数据年份",
            price_dir_labels,
            index=default_price_dir_index,
            help='读取"电价数据"目录下对应年份文件夹中的站点电价表。',
        )
        selected_price_dir = dict(price_dir_options)[selected_price_dir_label]

        # 扫描电价文件
        price_files = scan_price_files(selected_price_dir)

        if not price_files:
            st.warning("未找到电价数据文件！")
            st.info(f"请在以下目录放置电价文件: `{selected_price_dir}`")
            st.markdown("""
            **文件格式要求：**
            - Excel格式 (.xlsx/.xls)
            - 文件放在 `电价数据/年份` 文件夹内，例如 `电价数据/2025`
            - 第一列为日期列
            - 后续96列为00:00到23:45的96个时间节点电价数据
            """)
            return
        
        st.session_state['data_source'] = 'excel'

    has_busbar_info = (
        station_info_df is not None
        and '母线' in station_info_df.columns
        and '电站名' in station_info_df.columns
    )
    has_city_info = (
        station_info_df is not None
        and '城市' in station_info_df.columns
        and '电站名' in station_info_df.columns
    )
    busbar_types = []
    city_types = []

    if has_busbar_info:
        busbar_types = get_available_group_values(station_info_df, '母线')

    if has_city_info:
        # 直接从电站名表格获取所有城市，不按电价文件过滤
        city_types = (
            station_info_df['城市']
            .dropna()
            .astype(str)
            .str.strip()
            .replace('', pd.NA)
            .dropna()
            .unique()
            .tolist()
        )
        city_types = sorted(city_types)

    factory_group_types = get_available_group_values(
        station_info_df,
        FACTORY_GROUP_COLUMN,
        station_names=price_files.keys()
    )

    st.divider()
    if view_mode == "📊 电价数据查询":
        st.sidebar.header("📍 站点选择")

        # 级联筛选：城市 → 厂站类型 → 母线 → 站点
        filtered_stations = list(price_files.keys())

        # 1. 选择城市
        if city_types:
            city_options = ["全部城市"] + sorted(city_types)
            selected_city = st.sidebar.selectbox(
                "选择城市",
                options=city_options,
                index=0,
                help="选择要查看的城市"
            )
            
            if selected_city != "全部城市":
                station_group_df = build_station_group_mapping(
                    station_info_df,
                    "城市",
                    station_names=price_files.keys()
                )
                if station_group_df is not None:
                    city_stations_df = station_group_df[station_group_df["城市"] == selected_city]
                    city_stations_list = (
                        city_stations_df['电站名']
                        .astype(str)
                        .str.strip()
                        .tolist()
                    )
                    filtered_stations = [s for s in city_stations_list if s in price_files.keys()]
                    st.sidebar.info(f"📊 {selected_city} 共有 {len(filtered_stations)} 个站点")

        # 2. 在筛选后的站点中选择厂站类型
        if factory_group_types and len(filtered_stations) > 0:
            factory_group_options = get_available_group_values(
                station_info_df,
                FACTORY_GROUP_COLUMN,
                station_names=filtered_stations
            )
            
            if factory_group_options:
                all_factory_option = f"全部{FACTORY_GROUP_COLUMN}"
                factory_options = [all_factory_option] + factory_group_options
                selected_factory = st.sidebar.selectbox(
                    f"选择{FACTORY_GROUP_COLUMN}",
                    options=factory_options,
                    index=0,
                    help=f"选择要查看的{FACTORY_GROUP_COLUMN}"
                )
                
                if selected_factory != all_factory_option:
                    station_group_df = build_station_group_mapping(
                        station_info_df,
                        FACTORY_GROUP_COLUMN,
                        station_names=filtered_stations
                    )
                    if station_group_df is not None:
                        factory_stations_df = station_group_df[station_group_df[FACTORY_GROUP_COLUMN] == selected_factory]
                        factory_stations_list = (
                            factory_stations_df['电站名']
                            .astype(str)
                            .str.strip()
                            .tolist()
                        )
                        filtered_stations = [s for s in factory_stations_list if s in filtered_stations]
                        st.sidebar.info(f"📊 {selected_factory} 共有 {len(filtered_stations)} 个站点")

        # 3. 在筛选后的站点中选择母线
        if busbar_types and len(filtered_stations) > 0:
            busbar_options_filtered = get_available_group_values(
                station_info_df,
                "母线",
                station_names=filtered_stations
            )
            
            if busbar_options_filtered:
                all_busbar_option = "全部母线"
                busbar_options = [all_busbar_option] + busbar_options_filtered
                selected_busbar = st.sidebar.selectbox(
                    "选择母线",
                    options=busbar_options,
                    index=0,
                    help="选择要查看的母线"
                )
                
                if selected_busbar != all_busbar_option:
                    station_group_df = build_station_group_mapping(
                        station_info_df,
                        "母线",
                        station_names=filtered_stations
                    )
                    if station_group_df is not None:
                        busbar_stations_df = station_group_df[station_group_df["母线"] == selected_busbar]
                        busbar_stations_list = (
                            busbar_stations_df['电站名']
                            .astype(str)
                            .str.strip()
                            .tolist()
                        )
                        filtered_stations = [s for s in busbar_stations_list if s in filtered_stations]
                        st.sidebar.info(f" {selected_busbar} 共有 {len(filtered_stations)} 个站点")

        # 4. 最终站点选择
        if len(filtered_stations) == 0:
            st.warning("当前筛选条件下没有可用的电价数据文件！")
            return

        selected_station = st.sidebar.selectbox(
            "选择站点",
            options=filtered_stations,
            index=0,
            help="选择要查看的站点"
        )

        selected_file = price_files[selected_station]

        # 加载数据
        data_source = st.session_state.get('data_source', 'excel')
        
        with st.spinner("正在加载数据..."):
            if data_source == 'database':
                # 从数据库加载数据
                selected_year = st.session_state.get('selected_year')
                df = load_price_data_from_db(selected_station, selected_year)
            else:
                # 从Excel文件加载数据
                df = load_price_data(*get_file_cache_key(selected_file))

        if df is None:
            st.error("数据加载失败！")
            return

        # 计算全省平均电价（带磁盘缓存）
        if data_source == 'database':
            # 数据库模式：直接计算全省平均
            selected_year = st.session_state.get('selected_year')
            all_db_keys = [(station, selected_year) for station in price_files.keys()]
            prov_dates, prov_time_cols, prov_avg, failed_stations = compute_provincial_average_from_db(all_db_keys)
            
            # 显示失败站点提示和重新加载按钮
            if failed_stations:
                st.warning(f"以下 {len(failed_stations)} 个站点数据加载失败，已跳过: {', '.join(failed_stations[:5])}{'...' if len(failed_stations) > 5 else ''}")
                
                col_reload, col_info = st.columns([1, 3])
                with col_reload:
                    if st.button("🔄 重新加载失败站点", key="reload_failed_stations"):
                        # 清除相关缓存
                        compute_provincial_average_from_db.clear()
                        for station in failed_stations:
                            load_price_data_from_db.clear(station, selected_year)
                        st.rerun()
                with col_info:
                    st.info("点击按钮重新加载失败的站点数据")
        else:
            # Excel模式：使用原有缓存逻辑
            all_file_keys = [get_file_cache_key(fp) for fp in price_files.values()]
            prov_dates, prov_time_cols, prov_avg = compute_provincial_average_with_cache(tuple(all_file_keys))

        # 显示基本信息
        station_city = get_station_group_value(station_info_df, selected_station, '城市')
        station_title = f"{station_city} - {selected_station}" if station_city else selected_station
        st.header(f"📊 {station_title} - 电价数据概览")

        col1, col2, col3, col4 = st.columns(4)

        # 解析日期列
        date_col = df.columns[0]

        # 确保日期列是datetime类型
        try:
            df[date_col] = pd.to_datetime(df[date_col])
        except:
            pass

        # 计算统计信息
        price_cols = df.columns[1:]  # 除了日期列外的所有列都是电价数据
        all_prices = df[price_cols].values.flatten()

        with col1:
            st.metric("数据天数 (天)", f"{len(df)}")
        with col2:
            st.metric("最低电价 (元/kWh)", f"{all_prices.min():.4f}")
        with col3:
            st.metric("最高电价 (元/kWh)", f"{all_prices.max():.4f}")
        with col4:
            st.metric("平均电价 (元/kWh)", f"{all_prices.mean():.4f}")

        st.divider()

        # 日期选择器
        st.subheader("📅 选择日期查看详细数据")

        # 创建日期选项列表
        if pd.api.types.is_datetime64_any_dtype(df[date_col]):
            date_options = df[date_col].dt.strftime('%Y-%m-%d').tolist()
        else:
            date_options = df[date_col].astype(str).tolist()

        selected_date_idx = st.selectbox(
            "选择日期",
            options=range(len(date_options)),
            format_func=lambda x: date_options[x],
            index=0
        )

        # 获取选中日期的数据
        selected_row = df.iloc[selected_date_idx]
        selected_date = date_options[selected_date_idx]

        # 提取96个时间节点的电价
        time_columns = price_cols.tolist()
        prices = selected_row[price_cols].values.astype(float)

        # 显示选中日期的电价表格
        st.subheader(f" {selected_date} 电价明细表")

        # 创建展示用的DataFrame
        display_df = pd.DataFrame({
            '时间节点': time_columns,
            '电价 (元/kWh)': [round(p, 4) for p in prices]
        })

        # 添加时段类型标注
        selected_date_value = selected_row[date_col]
        display_df['时段类型'] = display_df['时间节点'].apply(
            lambda time_str: get_guangdong_period_type(selected_date_value, time_str)
        )

        # 显示表格（不使用样式）
        st.dataframe(display_df, use_container_width=True, height=400, column_config={
            "时间节点": st.column_config.TextColumn("时间节点", width="small"),
            "电价 (元/kWh)": st.column_config.TextColumn("电价 (元/kWh)", width="medium"),
            "时段类型": st.column_config.TextColumn("时段类型", width="small")
        }, hide_index=True)

        # 全省平均电价明细表
        if prov_avg is not None and prov_dates is not None:
            prov_date_to_row = {d: i for i, d in enumerate(prov_dates)}
            if selected_date in prov_date_to_row:
                prov_prices = prov_avg[prov_date_to_row[selected_date]]
                st.markdown("---")
                st.subheader(" 全省平均电价明细表")
                prov_display_df = pd.DataFrame({
                    '时间节点': time_columns,
                    '全省平均电价 (元/kWh)': np.round(prov_prices, 4)
                })
                prov_display_df['时段类型'] = prov_display_df['时间节点'].apply(
                    lambda time_str: get_guangdong_period_type(selected_date_value, time_str)
                )
                st.dataframe(prov_display_df, use_container_width=True, height=400, column_config={
                    "时间节点": st.column_config.TextColumn("时间节点", width="small"),
                    "全省平均电价 (元/kWh)": st.column_config.TextColumn("全省平均电价 (元/kWh)", width="medium"),
                    "时段类型": st.column_config.TextColumn("时段类型", width="small")
                }, hide_index=True)

        st.divider()

        # 可视化 - 电价曲线图
        st.subheader(" 电价曲线图")
        
        # 创建左右两列布局
        chart_col, metrics_col = st.columns([3, 1])
        
        with chart_col:
            # 创建Plotly图表
            fig = go.Figure()

            # 添加电价曲线
            fig.add_trace(go.Scatter(
                x=time_columns,
                y=prices,
                mode='lines+markers',
                name='电价',
                line=dict(color='#FF6B35', width=2),
                marker=dict(size=4)
            ))

            # 添加填充区域
            fig.add_trace(go.Scatter(
                x=time_columns,
                y=prices,
                mode='none',
                fill='tozeroy',
                fillcolor='rgba(255, 107, 53, 0.1)',
                name='电价区域'
            ))

            # 添加全省平均电价曲线
            if prov_avg is not None and prov_dates is not None:
                prov_date_to_row = {d: i for i, d in enumerate(prov_dates)}
                if selected_date in prov_date_to_row:
                    prov_prices = prov_avg[prov_date_to_row[selected_date]]
                    fig.add_trace(go.Scatter(
                        x=time_columns,
                        y=prov_prices,
                        mode='lines',
                        name='全省平均电价',
                        line=dict(color='blue', width=2, dash='dash')
                    ))

            # 更新布局
            fig.update_layout(
                title=f'{selected_date} 电价变化曲线（{selected_station} vs 全省平均）',
                xaxis_title='时间节点',
                yaxis_title='电价 (元/kWh)',
                hovermode='x unified',
                template='plotly_white',
                height=500,
                xaxis=dict(
                    tickangle=45,
                    tickvals=time_columns[::4],  # 每4个点显示一个标签
                    ticktext=[time_columns[i] for i in range(0, len(time_columns), 4)]
                )
            )

            st.plotly_chart(fig, use_container_width=True)
        
        with metrics_col:
            # 计算并显示站点电价与全省平均电价的标准差之和
            if prov_avg is not None and prov_dates is not None:
                prov_date_to_row = {d: i for i, d in enumerate(prov_dates)}
                if selected_date in prov_date_to_row:
                    prov_prices = prov_avg[prov_date_to_row[selected_date]]
                    
                    # 计算每个时间节点的标准差
                    price_diff = np.abs(prices - prov_prices)
                    std_sum = np.sum(price_diff)
                    
                    st.subheader("📊 电价差异分析")
                    st.divider()
                    
                    st.metric(
                        "标准差之和",
                        f"{std_sum:.4f}",
                        help="站点电价与全省平均电价在每个时间节点的绝对差值之和"
                    )
                    
                    st.metric(
                        "平均标准差",
                        f"{std_sum / len(prices):.4f}",
                        help="标准差之和的平均值"
                    )
                    
                    max_diff_idx = np.argmax(price_diff)
                    st.metric(
                        "最大差异节点",
                        time_columns[max_diff_idx],
                        help=f"差异最大的时间节点（差值：{price_diff[max_diff_idx]:.4f} 元/kWh）"
                    )
                    
                    # 添加额外信息
                    st.divider()
                    st.markdown(f"**时间节点数：** {len(prices)}")
                    st.markdown(f"**最小差值：** {np.min(price_diff):.4f}")
                    st.markdown(f"**最大差值：** {np.max(price_diff):.4f}")

        st.divider()

        # 价差规律分析
        if pd.api.types.is_datetime64_any_dtype(df[date_col]):
            st.subheader("🔍 价差规律分析")
            
            # 计算每日价差
            df_price_spread = df.copy()
            df_price_spread['日期'] = df_price_spread[date_col]
            df_price_spread['日最高电价'] = df_price_spread[price_cols].max(axis=1)
            df_price_spread['日最低电价'] = df_price_spread[price_cols].min(axis=1)
            df_price_spread['日电价差'] = df_price_spread['日最高电价'] - df_price_spread['日最低电价']
            df_price_spread['日均电价'] = df_price_spread[price_cols].mean(axis=1)
            df_price_spread['月份'] = df_price_spread['日期'].dt.month
            df_price_spread['季度'] = df_price_spread['日期'].dt.quarter
            df_price_spread['星期'] = df_price_spread['日期'].dt.dayofweek  # 0=周一, 6=周日
            df_price_spread['是否周末'] = df_price_spread['星期'].apply(lambda x: '周末' if x >= 5 else '工作日')
            
            # 创建Tabs
            tab1, tab2, tab3, tab4 = st.tabs(["日内价差规律", "工作日/周末对比", "月度价差趋势", "季度价差对比"])
            
            with tab1:
                st.markdown("### 📊 日内电价差分布分析")
                st.info("💡 **日内电价差** = 每日最高电价 - 每日最低电价，反映每天电价的最大波动幅度")
                
                # 获取日电价差数据
                daily_spread = df_price_spread['日电价差']
                
                # 统计信息
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("平均日电价差", f"{daily_spread.mean():.4f} 元/kWh")
                with col2:
                    st.metric("最大日电价差", f"{daily_spread.max():.4f} 元/kWh")
                with col3:
                    st.metric("最小日电价差", f"{daily_spread.min():.4f} 元/kWh")
                with col4:
                    st.metric("价差标准差", f"{daily_spread.std():.4f}")
                
                # 绘制日电价差分布直方图
                fig_hist = go.Figure()
                
                fig_hist.add_trace(go.Histogram(
                    x=daily_spread,
                    nbinsx=50,
                    name='日电价差分布',
                    marker_color='#FF6B35',
                    opacity=0.7
                ))
                
                # 添加平均线
                fig_hist.add_vline(
                    x=daily_spread.mean(),
                    line_dash="dash",
                    line_color="red",
                    annotation_text=f"平均值: {daily_spread.mean():.4f}",
                    annotation_position="top right"
                )
                
                fig_hist.update_layout(
                    title='日电价差分布直方图',
                    xaxis_title='日电价差 (元/kWh)',
                    yaxis_title='天数',
                    template='plotly_white',
                    height=400
                )
                
                st.plotly_chart(fig_hist, use_container_width=True)
                
                # 绘制箱线图
                st.markdown("### 📦 日电价差箱线图")
                
                fig_box = go.Figure()
                
                fig_box.add_trace(go.Box(
                    y=daily_spread,
                    name='日电价差',
                    marker_color='#FF6B35',
                    boxmean=True,
                    boxpoints='outliers'
                ))
                
                fig_box.update_layout(
                    title='日电价差箱线图（显示异常值）',
                    yaxis_title='日电价差 (元/kWh)',
                    template='plotly_white',
                    height=400
                )
                
                st.plotly_chart(fig_box, use_container_width=True)
                
                # 分位数分析
                st.markdown("### 📊 价差分位数统计")
                
                quantiles = [0.05, 0.10, 0.25, 0.50, 0.75, 0.90, 0.95]
                quantile_values = daily_spread.quantile(quantiles)
                
                quantile_df = pd.DataFrame({
                    '分位数': [f"{int(q*100)}%" for q in quantiles],
                    '价差值 (元/kWh)': quantile_values.round(4).values,
                    '说明': [
                        '仅5%的天数价差低于此值',
                        '仅10%的天数价差低于此值',
                        '25%的天数价差低于此值（下四分位数）',
                        '中位数（50%的天数价差低于此值）',
                        '75%的天数价差低于此值（上四分位数）',
                        '90%的天数价差低于此值',
                        '仅5%的天数价差高于此值'
                    ]
                })
                
                st.dataframe(quantile_df, use_container_width=True)
                
                # 关键发现
                st.markdown("### 🔍 关键发现")
                
                iqr = quantile_values[0.75] - quantile_values[0.25]
                lower_bound = quantile_values[0.25] - 1.5 * iqr
                upper_bound = quantile_values[0.75] + 1.5 * iqr
                
                outlier_count = ((daily_spread < lower_bound) | (daily_spread > upper_bound)).sum()
                
                col1, col2 = st.columns(2)
                with col1:
                    st.success(f" **中位数价差：** {quantile_values[0.50]:.4f} 元/kWh")
                    st.info(f" **75%的天数价差低于：** {quantile_values[0.75]:.4f} 元/kWh")
                with col2:
                    st.warning(f" **异常值天数：** {outlier_count} 天（占总天数的 {outlier_count/len(daily_spread)*100:.1f}%）")
                    st.markdown(f"- **正常价差范围：** {lower_bound:.4f} ~ {upper_bound:.4f} 元/kWh")
            
            with tab2:
                st.markdown("### 📊 工作日 vs 周末价差对比")
                
                # 按工作日/周末分组统计
                weekday_stats = df_price_spread.groupby('是否周末').agg({
                    '日电价差': ['mean', 'std', 'min', 'max', 'count']
                }).round(4)
                weekday_stats.columns = ['平均价差', '价差标准差', '最小价差', '最大价差', '天数']
                weekday_stats = weekday_stats.reset_index()
                
                # 显示统计表格
                st.dataframe(weekday_stats, use_container_width=True)
                
                # 绘制对比图
                fig_weekday = go.Figure()
                
                fig_weekday.add_trace(go.Bar(
                    x=weekday_stats['是否周末'],
                    y=weekday_stats['平均价差'],
                    name='平均价差',
                    marker_color=['#2196F3', '#FF6B35'],
                    text=weekday_stats['平均价差'].round(4),
                    textposition='auto'
                ))
                
                fig_weekday.update_layout(
                    title='工作日与周末平均价差对比',
                    xaxis_title='日期类型',
                    yaxis_title='平均价差 (元/kWh)',
                    template='plotly_white',
                    height=400
                )
                
                st.plotly_chart(fig_weekday, use_container_width=True)
            
            with tab3:
                st.markdown("### 📊 月度价差趋势")
                
                # 按月份统计价差
                monthly_spread = df_price_spread.groupby('月份').agg({
                    '日电价差': ['mean', 'std', 'min', 'max', 'count']
                }).round(4)
                monthly_spread.columns = ['平均价差', '价差标准差', '最小价差', '最大价差', '天数']
                monthly_spread = monthly_spread.reset_index()
                
                # 显示统计表格
                st.dataframe(monthly_spread, use_container_width=True)
                
                # 绘制月度趋势图
                fig_monthly = go.Figure()
                
                fig_monthly.add_trace(go.Scatter(
                    x=monthly_spread['月份'],
                    y=monthly_spread['平均价差'],
                    mode='lines+markers',
                    name='平均价差',
                    line=dict(color='#FF6B35', width=2),
                    marker=dict(size=8)
                ))
                
                # 添加误差线（标准差）
                fig_monthly.add_trace(go.Scatter(
                    x=monthly_spread['月份'],
                    y=monthly_spread['平均价差'] + monthly_spread['价差标准差'],
                    mode='lines',
                    name='+标准差',
                    line=dict(color='rgba(255, 107, 53, 0.3)', width=1, dash='dash'),
                    showlegend=False
                ))
                
                fig_monthly.add_trace(go.Scatter(
                    x=monthly_spread['月份'],
                    y=monthly_spread['平均价差'] - monthly_spread['价差标准差'],
                    mode='lines',
                    name='-标准差',
                    line=dict(color='rgba(255, 107, 53, 0.3)', width=1, dash='dash'),
                    fill='tonexty',
                    fillcolor='rgba(255, 107, 53, 0.1)',
                    showlegend=False
                ))
                
                fig_monthly.update_layout(
                    title='月度平均价差趋势',
                    xaxis_title='月份',
                    yaxis_title='平均价差 (元/kWh)',
                    template='plotly_white',
                    height=400,
                    xaxis=dict(tickmode='linear', dtick=1)
                )
                
                st.plotly_chart(fig_monthly, use_container_width=True)
            
            with tab4:
                st.markdown("### 📊 季度价差对比")
                
                # 按季度统计价差
                quarterly_spread = df_price_spread.groupby('季度').agg({
                    '日电价差': ['mean', 'std', 'min', 'max', 'count']
                }).round(4)
                quarterly_spread.columns = ['平均价差', '价差标准差', '最小价差', '最大价差', '天数']
                quarterly_spread = quarterly_spread.reset_index()
                quarterly_spread['季度名称'] = quarterly_spread['季度'].apply(lambda x: f'Q{x}')
                
                # 显示统计表格
                st.dataframe(quarterly_spread[['季度名称', '平均价差', '价差标准差', '最小价差', '最大价差', '天数']], use_container_width=True)
                
                # 绘制季度对比图
                fig_quarterly = go.Figure()
                
                fig_quarterly.add_trace(go.Bar(
                    x=quarterly_spread['季度名称'],
                    y=quarterly_spread['平均价差'],
                    name='平均价差',
                    marker_color=['#FF6B35', '#2196F3', '#4CAF50', '#9C27B0'],
                    text=quarterly_spread['平均价差'].round(4),
                    textposition='auto'
                ))
                
                fig_quarterly.update_layout(
                    title='季度平均价差对比',
                    xaxis_title='季度',
                    yaxis_title='平均价差 (元/kWh)',
                    template='plotly_white',
                    height=400
                )
                
                st.plotly_chart(fig_quarterly, use_container_width=True)
        
        st.divider()

        # 时段特征及波动趋势分析
        if pd.api.types.is_datetime64_any_dtype(df[date_col]):
            st.subheader("⚡ 时段特征及波动趋势")
            
            # 为每个时间节点添加时段类型标签
            df_time_features = df.copy()
            df_time_features['日期'] = df_time_features[date_col]
            
            # 计算每个时间节点的统计特征
            time_stats_list = []
            for time_col in time_columns:
                time_data = df_time_features[time_col].dropna()
                time_stats = {
                    '时间节点': time_col,
                    '平均电价': time_data.mean(),
                    '电价中位数': time_data.median(),
                    '电价标准差': time_data.std(),
                    '电价变异系数': time_data.std() / time_data.mean() if time_data.mean() != 0 else 0,
                    '最高电价': time_data.max(),
                    '最低电价': time_data.min(),
                    '电价极差': time_data.max() - time_data.min(),
                    '数据天数': len(time_data)
                }
                time_stats_list.append(time_stats)
            
            time_stats_df = pd.DataFrame(time_stats_list).round(4)
            
            # 添加时段类型
            # 假设第一个日期来判定时段类型
            sample_date = df_time_features[date_col].iloc[0]
            time_stats_df['时段类型'] = time_stats_df['时间节点'].apply(
                lambda t: get_guangdong_period_type(sample_date, t)
            )
            
            # 创建Tabs
            tab1, tab2, tab3 = st.tabs(["峰平谷时段特征", "电价波动趋势", "波动特征统计"])
            
            with tab1:
                st.markdown("###  峰、平、谷时段电价特征")
                
                # 按时段类型分组统计
                period_stats = time_stats_df.groupby('时段类型').agg({
                    '平均电价': 'mean',
                    '电价标准差': 'mean',
                    '电价变异系数': 'mean',
                    '最高电价': 'max',
                    '最低电价': 'min',
                    '电价极差': 'mean'
                }).round(4).reset_index()
                
                # 显示统计表格
                st.dataframe(period_stats, use_container_width=True)
                
                # 绘制峰平谷对比图
                fig_period = go.Figure()
                
                # 按顺序排列时段类型
                period_order = ['尖峰', '峰', '平', '谷', '深谷']
                period_stats['排序'] = period_stats['时段类型'].apply(
                    lambda x: period_order.index(x) if x in period_order else len(period_order)
                )
                period_stats = period_stats.sort_values('排序')
                
                colors = {'尖峰': '#FF0000', '峰': '#FF6B35', '平': '#FFC107', '谷': '#2196F3', '深谷': '#4CAF50'}
                
                fig_period.add_trace(go.Bar(
                    x=period_stats['时段类型'],
                    y=period_stats['平均电价'],
                    name='平均电价',
                    marker_color=[colors.get(p, '#999999') for p in period_stats['时段类型']],
                    text=period_stats['平均电价'].round(4),
                    textposition='auto'
                ))
                
                fig_period.update_layout(
                    title='各时段类型平均电价对比',
                    xaxis_title='时段类型',
                    yaxis_title='平均电价 (元/kWh)',
                    template='plotly_white',
                    height=400
                )
                
                st.plotly_chart(fig_period, use_container_width=True)
                
                # 显示关键发现
                col1, col2, col3 = st.columns(3)
                with col1:
                    max_period = period_stats.loc[period_stats['平均电价'].idxmax()]
                    st.success(f" **最高均价时段：** {max_period['时段类型']} ({max_period['平均电价']:.4f} 元/kWh)")
                with col2:
                    min_period = period_stats.loc[period_stats['平均电价'].idxmin()]
                    st.info(f"🟢 **最低均价时段：** {min_period['时段类型']} ({min_period['平均电价']:.4f} 元/kWh)")
                with col3:
                    max_spread_period = period_stats.loc[period_stats['电价极差'].idxmax()]
                    st.warning(f"🟡 **波动最大时段：** {max_spread_period['时段类型']} (极差：{max_spread_period['电价极差']:.4f})")
            
            with tab2:
                st.markdown("### 📊 各时间节点电价波动趋势")
                
                # 绘制每个时间节点的均价和标准差
                fig_trend = go.Figure()
                
                # 均价曲线
                fig_trend.add_trace(go.Scatter(
                    x=time_stats_df['时间节点'],
                    y=time_stats_df['平均电价'],
                    mode='lines+markers',
                    name='平均电价',
                    line=dict(color='#FF6B35', width=2),
                    marker=dict(size=6)
                ))
                
                # 标准差带
                fig_trend.add_trace(go.Scatter(
                    x=time_stats_df['时间节点'],
                    y=time_stats_df['平均电价'] + time_stats_df['电价标准差'],
                    mode='lines',
                    name='+标准差',
                    line=dict(color='rgba(255, 107, 53, 0.3)', width=1, dash='dash'),
                    showlegend=False
                ))
                
                fig_trend.add_trace(go.Scatter(
                    x=time_stats_df['时间节点'],
                    y=time_stats_df['平均电价'] - time_stats_df['电价标准差'],
                    mode='lines',
                    name='-标准差',
                    line=dict(color='rgba(255, 107, 53, 0.3)', width=1, dash='dash'),
                    fill='tonexty',
                    fillcolor='rgba(255, 107, 53, 0.1)',
                    showlegend=False
                ))
                
                fig_trend.update_layout(
                    title='各时间节点平均电价及波动范围',
                    xaxis_title='时间节点',
                    yaxis_title='电价 (元/kWh)',
                    template='plotly_white',
                    height=400,
                    xaxis=dict(
                        tickangle=45,
                        tickvals=time_columns[::4],
                        ticktext=[time_columns[i] for i in range(0, len(time_columns), 4)]
                    )
                )
                
                st.plotly_chart(fig_trend, use_container_width=True)
                
                # 变异系数趋势图
                fig_cv = go.Figure()
                
                fig_cv.add_trace(go.Scatter(
                    x=time_stats_df['时间节点'],
                    y=time_stats_df['电价变异系数'],
                    mode='lines+markers',
                    name='变异系数',
                    line=dict(color='#2196F3', width=2),
                    marker=dict(size=6)
                ))
                
                fig_cv.update_layout(
                    title='各时间节点电价变异系数（相对波动程度）',
                    xaxis_title='时间节点',
                    yaxis_title='变异系数',
                    template='plotly_white',
                    height=400,
                    xaxis=dict(
                        tickangle=45,
                        tickvals=time_columns[::4],
                        ticktext=[time_columns[i] for i in range(0, len(time_columns), 4)]
                    )
                )
                
                st.plotly_chart(fig_cv, use_container_width=True)
            
            with tab3:
                st.markdown("###  波动特征统计")
                
                # 显示详细的时间节点统计数据
                st.dataframe(time_stats_df, use_container_width=True, height=400)
                
                # 找出波动最大和最小的时间段
                col1, col2 = st.columns(2)
                with col1:
                    max_std_row = time_stats_df.loc[time_stats_df['电价标准差'].idxmax()]
                    st.warning(f" **波动最大的时间节点：** {max_std_row['时间节点']}")
                    st.markdown(f"- 平均电价：{max_std_row['平均电价']:.4f} 元/kWh")
                    st.markdown(f"- 标准差：{max_std_row['电价标准差']:.4f}")
                    st.markdown(f"- 变异系数：{max_std_row['电价变异系数']:.4f}")
                    st.markdown(f"- 极差：{max_std_row['电价极差']:.4f}")
                
                with col2:
                    min_std_row = time_stats_df.loc[time_stats_df['电价标准差'].idxmin()]
                    st.success(f" **波动最小的时间节点：** {min_std_row['时间节点']}")
                    st.markdown(f"- 平均电价：{min_std_row['平均电价']:.4f} 元/kWh")
                    st.markdown(f"- 标准差：{min_std_row['电价标准差']:.4f}")
                    st.markdown(f"- 变异系数：{min_std_row['电价变异系数']:.4f}")
                    st.markdown(f"- 极差：{min_std_row['电价极差']:.4f}")
        
        st.divider()

        # 年度统计
        st.subheader(" 年度统计分析")

        # 计算每天的平均电价
        daily_avg_prices = df[price_cols].mean(axis=1)

        # 添加日期列
        if pd.api.types.is_datetime64_any_dtype(df[date_col]):
            analysis_df = pd.DataFrame({
                '日期': df[date_col].dt.strftime('%Y-%m-%d'),
                '日均电价': daily_avg_prices.values
            })
        else:
            analysis_df = pd.DataFrame({
                '日期': df[date_col].astype(str),
                '日均电价': daily_avg_prices.values
            })

        # 绘制年度电价趋势图
        fig_trend = go.Figure()

        fig_trend.add_trace(go.Scatter(
            x=analysis_df['日期'],
            y=analysis_df['日均电价'],
            mode='lines',
            name=f'{selected_station} 日均电价',
            line=dict(color='#2196F3', width=1.5)
        ))

        # 添加全省平均日均电价曲线
        if prov_avg is not None and prov_dates is not None:
            prov_daily_avg = prov_avg.mean(axis=1)
            fig_trend.add_trace(go.Scatter(
                x=prov_dates,
                y=prov_daily_avg,
                mode='lines',
                name='全省平均日均电价',
                line=dict(color='red', width=1.5, dash='dash')
            ))

        fig_trend.update_layout(
            title=f'{selected_station} vs 全省平均 日均电价趋势对比',
            xaxis_title='日期',
            yaxis_title='日均电价 (元/kWh)',
            template='plotly_white',
            height=400,
            xaxis=dict(
                tickangle=45,
                tickvals=analysis_df['日期'][::30],  # 每30天显示一个标签
                ticktext=[analysis_df['日期'].iloc[i] for i in range(0, len(analysis_df), 30)]
            )
        )

        st.plotly_chart(fig_trend, use_container_width=True)

        # 月度统计
        st.subheader("📅 月度统计摘要")

        if pd.api.types.is_datetime64_any_dtype(df[date_col]):
            # 提取月份
            df_copy = df.copy()
            df_copy['月份'] = df_copy[date_col].dt.to_period('M').astype(str)

            # 计算每天的平均电价
            df_copy['日均电价'] = df_copy[price_cols].mean(axis=1)

            # 按月统计
            monthly_stats = df_copy.groupby('月份')['日均电价'].agg(['mean', 'min', 'max', 'std'])
            monthly_stats.columns = ['月均电价', '月最低电价', '月最高电价', '电价标准差']
            monthly_stats = monthly_stats.round(4)

            st.dataframe(monthly_stats, use_container_width=True)
        else:
            st.info("无法进行月度统计（日期格式不支持）")

        st.divider()

        # 数据导出
        st.subheader("💾 数据导出")

        export_option = st.radio(
            "选择导出内容",
            ["单日详细数据", "全年完整数据"],
            horizontal=True
        )

        if export_option == "单日详细数据":
            # 重新创建单日数据
            single_day_df = pd.DataFrame({
                '时间节点': time_columns,
                '电价 (元/kWh)': prices
            })
            single_day_df.insert(0, '日期', selected_date)
            csv_data = single_day_df.to_csv(index=False, encoding='utf-8-sig')
            file_name = f"{selected_station}_{selected_date}_电价数据.csv"
        else:
            export_df = df.copy()
            csv_data = export_df.to_csv(index=False, encoding='utf-8-sig')
            file_name = f"{selected_station}_全年电价数据.csv"

        st.download_button(
            label="📥 下载CSV文件",
            data=csv_data,
            file_name=file_name,
            mime="text/csv"
        )

    # 电价差排名页
    elif view_mode == "📈 电价差排名":
        # 侧边栏：全局筛选
        selected_city_filter = None
        if city_types:
            st.sidebar.divider()
            st.sidebar.header("🌍 全局筛选")
            all_city_option = "全部城市"
            selected_city_filter = st.sidebar.selectbox(
                "选择城市",
                options=[all_city_option] + city_types,
                index=0,
                help="选择城市后，所有排名都将限制在该城市范围内"
            )
            if selected_city_filter == all_city_option:
                selected_city_filter = None
        
        # 如果选择了城市，先过滤站点
        filtered_price_files = price_files
        if selected_city_filter:
            city_station_df = build_station_group_mapping(
                station_info_df,
                '城市',
                station_names=list(price_files.keys())
            )
            if city_station_df is not None:
                city_stations = city_station_df[
                    city_station_df['城市'] == selected_city_filter
                ]['电站名'].astype(str).str.strip().tolist()
                filtered_price_files = {k: v for k, v in price_files.items() if k in city_stations}
        
        # 根据数据源选择计算方式
        data_source = st.session_state.get('data_source', 'excel')
        
        if data_source == 'database':
            # 数据库模式
            selected_year = st.session_state.get('selected_year')
            station_year_pairs = [(station, selected_year) for station in filtered_price_files.keys()]
            st.caption("从数据库计算电价差排名。")
            
            if selected_city_filter:
                st.caption(f"🌍 当前筛选城市：{selected_city_filter}")

            with st.spinner("正在计算所有站点电价差..."):
                all_stations_stats, failed_stations, cache_summary = calculate_all_stations_price_spread_from_db(station_year_pairs)

            st.caption(
                f"本次从数据库计算 {cache_summary['recomputed_count']} 个站点。"
            )
            
            # 显示失败站点提示和重新加载按钮
            if failed_stations:
                st.warning(f"有 {len(failed_stations)} 个站点统计失败，已自动跳过: {', '.join([s[0] if isinstance(s, tuple) else s for s in failed_stations[:5]])}{'...' if len(failed_stations) > 5 else ''}")
                
                col_reload, col_info = st.columns([1, 3])
                with col_reload:
                    if st.button("🔄 重新加载失败站点", key="reload_failed_rankings"):
                        # 清除相关缓存
                        calculate_all_stations_price_spread_from_db.clear()
                        for station_info in failed_stations:
                            station_name = station_info[0] if isinstance(station_info, tuple) else station_info
                            load_price_data_from_db.clear(station_name, selected_year)
                        st.rerun()
                with col_info:
                    st.info("点击按钮重新加载失败的站点数据")
        else:
            # Excel模式
            price_file_index = build_price_file_index(filtered_price_files)
            st.caption("排名结果会自动缓存到本地，应用重启后也能复用；只有变更过的 Excel 才会重算。")
            
            if selected_city_filter:
                st.caption(f"🌍 当前筛选城市：{selected_city_filter}")

            with st.spinner("正在计算所有站点电价差..."):
                all_stations_stats, failed_stations, cache_summary = calculate_all_stations_price_spread(price_file_index)

            st.caption(
                f"本次命中本地缓存 {cache_summary['cached_count']} 个文件，"
                f"重新计算 {cache_summary['recomputed_count']} 个文件。"
            )
            
            # Excel模式下显示失败站点提示
            if failed_stations:
                st.warning(f"有 {len(failed_stations)} 个站点文件统计失败，已自动跳过。")

        if len(all_stations_stats) > 0:
            # 为所有统计数据添加厂站类型信息
            all_stations_with_factory = prepare_grouped_rankings(
                all_stations_stats, station_info_df, FACTORY_GROUP_COLUMN
            )
            
            grouped_ranking_configs = {}

            if busbar_types:
                grouped_busbar_df = prepare_grouped_rankings(all_stations_stats, station_info_df, '母线')
                if grouped_busbar_df is not None:
                    grouped_ranking_configs["按母线排名"] = {
                        "dataframe": grouped_busbar_df,
                        "group_column": "母线"
                    }

            ranking_mode_options = ["全部站点总排名"] + list(grouped_ranking_configs.keys())
            if not grouped_ranking_configs:
                st.info("未找到可用的分组信息，当前仅显示全部站点总排名。")

            ranking_mode = st.radio(
                "排名视图",
                options=ranking_mode_options,
                horizontal=True
            )

            if ranking_mode == "全部站点总排名":
                # 添加厂站类型筛选器
                factory_filter_options = ["全部"] + factory_group_types if factory_group_types else ["全部"]
                selected_factory_filter = st.radio(
                    "厂站类型筛选",
                    options=factory_filter_options,
                    horizontal=True,
                    index=0
                )
                
                # 根据厂站类型筛选数据
                if selected_factory_filter == "全部":
                    filtered_stats_df = all_stations_stats.copy()
                else:
                    filtered_stats_df = all_stations_with_factory[
                        all_stations_with_factory[FACTORY_GROUP_COLUMN] == selected_factory_filter
                    ].copy()
                
                if len(filtered_stats_df) == 0:
                    st.warning(f"当前筛选条件下没有{selected_factory_filter}的数据！")
                    return
                
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("总站点数", f"{len(filtered_stats_df)} 个")
                with col2:
                    max_spread_station = filtered_stats_df.iloc[0]
                    st.metric("最高电价差", f"{max_spread_station['日均电价差']:.4f} 元/kWh",
                             delta=max_spread_station['站点名称'])
                with col3:
                    min_spread_station = filtered_stats_df.iloc[-1]
                    st.metric("最低电价差", f"{min_spread_station['日均电价差']:.4f} 元/kWh",
                             delta=min_spread_station['站点名称'], delta_color="inverse")
                with col4:
                    avg_spread_all = filtered_stats_df['日均电价差'].mean()
                    st.metric("平均电价差", f"{avg_spread_all:.4f} 元/kWh")

                table_title = f"📋 {'全部站点' if selected_factory_filter == '全部' else selected_factory_filter}总排名明细表"
                display_df = filtered_stats_df.copy()
                chart_source_df = display_df.copy()
                chart_x_col = '站点名称'
                chart_title_prefix = "前"
                chart_title_suffix = "个站点日均电价差对比"
                chart_x_title = "站点名称"
                export_file_name = f"站点电价差总排名{'_' + selected_factory_filter if selected_factory_filter != '全部' else ''}.csv"
                chart_caption = "表格保留全部站点，图表默认只展示前若干名，避免一次性渲染过多柱子影响速度。"
                chart_hover_station = False
                chart_group_label = None
            else:
                grouped_config = grouped_ranking_configs[ranking_mode]
                group_column = grouped_config["group_column"]
                group_rank_column_label = f"{group_column}内排名"
                grouped_display_df = grouped_config["dataframe"].rename(columns={'排名': '总排名'}).copy()
                
                # 为分组排名数据添加厂站类型信息
                factory_group_df = build_station_group_mapping(
                    station_info_df,
                    FACTORY_GROUP_COLUMN,
                    station_names=grouped_display_df['站点名称'].tolist()
                )
                if factory_group_df is not None:
                    grouped_display_df = grouped_display_df.merge(
                        factory_group_df[['电站名', FACTORY_GROUP_COLUMN]],
                        left_on='站点名称',
                        right_on='电站名',
                        how='left'
                    ).drop(columns=['电站名'])
                    grouped_display_df[FACTORY_GROUP_COLUMN] = grouped_display_df[FACTORY_GROUP_COLUMN].fillna('未分组')
                
                # 添加厂站类型筛选器
                factory_filter_options = ["全部"] + factory_group_types if factory_group_types else ["全部"]
                selected_factory_filter = st.radio(
                    "厂站类型筛选",
                    options=factory_filter_options,
                    horizontal=True,
                    index=0
                )
                
                # 根据厂站类型筛选数据
                if selected_factory_filter != "全部":
                    grouped_display_df = grouped_display_df[
                        grouped_display_df[FACTORY_GROUP_COLUMN] == selected_factory_filter
                    ].copy()
                    # 重新计算组内排名
                    grouped_display_df = grouped_display_df.sort_values(
                        [group_column, '日均电价差', '站点名称'],
                        ascending=[True, False, True]
                    ).reset_index(drop=True)
                    grouped_display_df['组内排名'] = grouped_display_df.groupby(group_column).cumcount() + 1
                
                available_groups = sorted(grouped_display_df[group_column].unique().tolist())
                selected_rank_group = st.selectbox(
                    f"选择{group_column}",
                    options=[f"全部{group_column}"] + available_groups,
                    index=0
                )

                if selected_rank_group == f"全部{group_column}":
                    champion_df = (
                        grouped_display_df[grouped_display_df['组内排名'] == 1]
                        .sort_values('日均电价差', ascending=False)
                        .reset_index(drop=True)
                    )

                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric(f"{group_column}组数", f"{len(available_groups)} 个")
                    with col2:
                        st.metric("总站点数", f"{len(grouped_display_df)} 个")
                    with col3:
                        champion_station = champion_df.iloc[0]
                        st.metric(f"最强{group_column}第1名", f"{champion_station['日均电价差']:.4f} 元/kWh",
                                 delta=f"{champion_station[group_column]} - {champion_station['站点名称']}")
                    with col4:
                        st.metric("冠军平均电价差", f"{champion_df['日均电价差'].mean():.4f} 元/kWh")

                    table_title = f"📋 各{group_column}排名明细表"
                    display_df = grouped_display_df[
                        [group_column, '组内排名', '总排名', '站点名称', '日均电价差', '全年最高电价差', '全年最低电价差', '全年平均电价', '数据天数']
                    ].copy()
                    display_df = display_df.sort_values([group_column, '组内排名']).reset_index(drop=True)
                    display_df = display_df.rename(columns={'组内排名': group_rank_column_label})
                    chart_source_df = champion_df.copy()
                    chart_x_col = group_column
                    chart_title_prefix = "前"
                    chart_title_suffix = f"个{group_column}冠军电价差对比"
                    chart_x_title = group_column
                    export_file_name = f"{group_column}排名_全部{group_column}.csv"
                    chart_caption = f"表格显示全部{group_column}的排名，图表展示各{group_column}第1名的对比。"
                    chart_hover_station = True
                    chart_group_label = group_column
                else:
                    single_group_df = grouped_display_df[
                        grouped_display_df[group_column] == selected_rank_group
                    ].sort_values('组内排名').reset_index(drop=True)

                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric(f"当前{group_column}站点数", f"{len(single_group_df)} 个")
                    with col2:
                        top_station = single_group_df.iloc[0]
                        st.metric(f"{group_column}第1名", f"{top_station['日均电价差']:.4f} 元/kWh",
                                 delta=top_station['站点名称'])
                    with col3:
                        bottom_station = single_group_df.iloc[-1]
                        st.metric(f"{group_column}最后1名", f"{bottom_station['日均电价差']:.4f} 元/kWh",
                                 delta=bottom_station['站点名称'], delta_color="inverse")
                    with col4:
                        st.metric(f"当前{group_column}平均电价差", f"{single_group_df['日均电价差'].mean():.4f} 元/kWh")

                    table_title = f"📋 {selected_rank_group} 排名明细表"
                    display_df = single_group_df[
                        ['组内排名', '总排名', '站点名称', '日均电价差', '全年最高电价差', '全年最低电价差', '全年平均电价', '数据天数']
                    ].copy()
                    display_df = display_df.rename(columns={'组内排名': group_rank_column_label})
                    chart_source_df = single_group_df.copy()
                    chart_x_col = '站点名称'
                    chart_title_prefix = selected_rank_group
                    chart_title_suffix = "排名电价差对比"
                    chart_x_title = "站点名称"
                    export_file_name = f"{selected_rank_group}_{group_column}排名.csv"
                    chart_caption = f"表格和图表都按当前{group_column}展示排名。"
                    chart_hover_station = False
                    chart_group_label = None

            st.subheader(table_title)
            st.dataframe(display_df, use_container_width=True, height=400)

            st.subheader("📈 电价差分布图")
            st.caption(chart_caption)

            if len(chart_source_df) == 0:
                st.info("当前条件下没有可用于绘图的数据。")
            else:
                chart_options = []
                for option in [20, 50, 100, 200, len(chart_source_df)]:
                    if option <= len(chart_source_df) and option not in chart_options:
                        chart_options.append(option)

                if len(chart_options) == 1:
                    chart_station_count = chart_options[0]
                    st.caption(f"当前仅有 {chart_station_count} 条数据，图表已自动展示全部。")
                else:
                    default_chart_count = 50 if 50 in chart_options else chart_options[-1]
                    chart_station_count = st.select_slider(
                        "图表展示数量",
                        options=chart_options,
                        value=default_chart_count
                    )

                chart_df = chart_source_df.head(chart_station_count).copy()
                show_value_labels = chart_station_count <= 50

                fig_spread = go.Figure()
                bar_kwargs = dict(
                    x=chart_df[chart_x_col],
                    y=chart_df['日均电价差'],
                    name='日均电价差',
                    marker_color='rgb(55, 83, 109)'
                )
                if show_value_labels:
                    bar_kwargs['text'] = [f"{x:.4f}" for x in chart_df['日均电价差']]
                    bar_kwargs['textposition'] = 'auto'
                if chart_hover_station:
                    bar_kwargs['customdata'] = chart_df[['站点名称']].to_numpy()
                    bar_kwargs['hovertemplate'] = f"{chart_group_label}: %{{x}}<br>站点: %{{customdata[0]}}<br>日均电价差: %{{y:.4f}} 元/kWh<extra></extra>"

                fig_spread.add_trace(go.Bar(**bar_kwargs))

                if ranking_mode == "全部站点总排名" or chart_x_col != '站点名称':
                    chart_title = f"{chart_title_prefix} {chart_station_count} {chart_title_suffix}"
                else:
                    chart_title = f"{chart_title_prefix}{chart_title_suffix}"

                fig_spread.update_layout(
                    title=chart_title,
                    xaxis_title=chart_x_title,
                    yaxis_title='日均电价差 (元/kWh)',
                    template='plotly_white',
                    height=500,
                    xaxis=dict(tickangle=45),
                    showlegend=False
                )

                st.plotly_chart(fig_spread, use_container_width=True)

            st.subheader("💾 导出排名数据")
            csv_data = display_df.to_csv(index=False, encoding='utf-8-sig')
            st.download_button(
                label="📥 下载排名数据 (CSV)",
                data=csv_data,
                file_name=export_file_name,
                mime="text/csv"
            )
        else:
            st.warning("暂无电价差数据！")

    # 储能配储优化页
    elif view_mode == "🔋 储能配储优化":
        st.header(" 储能配储优化")
        st.markdown("""
        基于站点一年电价数据，进行储能配置优化，计算最优充放电策略和年收益。
        """)
            
        # 侧边栏：选择站点和储能参数
        st.sidebar.divider()
        st.sidebar.header("🔋 储能配储优化配置")
        
        # 站点筛选方式
        st.sidebar.subheader("📍 站点选择")
        filter_mode_options = ["全部站点"]
        if busbar_types:
            filter_mode_options.append("按母线查询")
        if city_types:
            filter_mode_options.append("按城市查询")
        if factory_group_types:
            filter_mode_options.append("按厂站类型查询")
        
        selected_filter_mode = st.sidebar.selectbox(
            "查询方式",
            options=filter_mode_options,
            index=0,
            help="选择站点筛选方式"
        )
        
        filtered_stations = list(price_files.keys())
        
        # 根据筛选方式过滤站点
        if selected_filter_mode == "按母线查询":
            group_options = busbar_types
            active_group_column = "母线"
        elif selected_filter_mode == "按城市查询":
            group_options = city_types
            active_group_column = "城市"
        elif selected_filter_mode == "按厂站类型查询":
            group_options = factory_group_types
            active_group_column = FACTORY_GROUP_COLUMN
        else:
            group_options = []
            active_group_column = None
        
        if active_group_column and group_options:
            all_group_option = f"全部{active_group_column}"
            active_group_value = st.sidebar.selectbox(
                f"选择{active_group_column}",
                options=[all_group_option] + group_options,
                index=0,
                help=f"选择要查看的{active_group_column}"
            )
            
            if active_group_value != all_group_option:
                station_group_df = build_station_group_mapping(
                    station_info_df,
                    active_group_column,
                    station_names=price_files.keys()
                )
                filtered_stations_df = station_group_df[
                    station_group_df[active_group_column] == active_group_value
                ]
                filtered_stations_list = (
                    filtered_stations_df['电站名']
                    .astype(str)
                    .str.strip()
                    .tolist()
                )
                filtered_stations = [s for s in filtered_stations_list if s in price_files.keys()]
                st.sidebar.info(f"📊 {active_group_value} 共有 {len(filtered_stations)} 个站点")
        
        if len(filtered_stations) == 0:
            st.sidebar.warning("当前筛选条件下没有可用的站点！")
            selected_station = None
        else:
            # 选择站点
            selected_station = st.sidebar.selectbox(
                "选择站点",
                options=filtered_stations,
                index=0,
                help="选择要进行储能优化的站点"
            )
            
        # 储能参数配置
        st.sidebar.subheader("⚙️ 储能参数")
        P = st.sidebar.number_input(
            "逆变器功率 (kW)",
            value=STORAGE_CONFIG['P'],
            min_value=1000,
            step=10000,
            help="储能逆变器额定功率"
        )
        battery_capacity = st.sidebar.number_input(
            "电池容量 (kWh)",
            value=STORAGE_CONFIG['battery_capacity'],
            min_value=1000,
            step=10000,
            help="电池总容量"
        )
        efficiency = st.sidebar.number_input(
            "放电效率",
            value=STORAGE_CONFIG['efficiency'],
            min_value=0.5,
            max_value=1.0,
            step=0.01,
            help="充放电循环效率"
        )
        
        # 初始化 session_state
        if "opt_cache_key" not in st.session_state:
            st.session_state.opt_cache_key = None
        if "opt_results" not in st.session_state:
            st.session_state.opt_results = None

        # 根据数据源生成缓存键
        data_source = st.session_state.get('data_source', 'excel')
        if data_source == 'database':
            selected_year = st.session_state.get('selected_year')
            cache_key = (f"db_{selected_station}_{selected_year}", P, battery_capacity, efficiency)
        else:
            cache_key = (str(price_files.get(selected_station, '')), P, battery_capacity, efficiency)
        
        params_changed = (st.session_state.opt_cache_key != cache_key)

        # 开始优化按钮
        run_opt = st.button("🚀 开始优化", type="primary", use_container_width=True)

        if run_opt:
            with st.spinner("正在加载电价数据并优化..."):
                try:
                    if data_source == 'database':
                        # 数据库模式
                        selected_year = st.session_state.get('selected_year')
                        date_list, all_results, all_summaries = run_optimization_cached_from_db(
                            selected_station, selected_year, P, battery_capacity, efficiency
                        )
                    else:
                        # Excel模式
                        date_list, all_results, all_summaries = run_optimization_cached(
                            price_files[selected_station], P, battery_capacity, efficiency
                        )
                    
                    st.session_state.opt_results = {
                        'date_list': date_list,
                        'all_results': all_results,
                        'all_summaries': all_summaries,
                    }
                    st.session_state.opt_cache_key = cache_key
                except Exception as e:
                    st.error(f"优化失败：{str(e)}")
                    st.exception(e)
                    st.stop()

        # 显示已缓存的结果（按钮按下后或之前已计算过且参数未变）
        if st.session_state.opt_results is not None and not params_changed:
            date_list = st.session_state.opt_results['date_list']
            all_results = st.session_state.opt_results['all_results']
            all_summaries = st.session_state.opt_results['all_summaries']
            num_days = len(all_summaries)

            # 计算年收益
            total_yearly_revenue = sum([s['日收益_元'] for s in all_summaries])

            # 显示年收益
            st.success(f"🎉 优化完成！该站点年收益：**{total_yearly_revenue:,.2f} 元**")

            # 显示参数和收益摘要
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("逆变器功率", f"{P/1000:.0f} MW")
            with col2:
                st.metric("电池容量", f"{battery_capacity/1000:.0f} MWh")
            with col3:
                st.metric("年收益", f"{total_yearly_revenue:,.0f} 元")
            with col4:
                st.metric("日均收益", f"{total_yearly_revenue/num_days:.0f} 元")

            # 每日收益折线图
            st.divider()
            st.subheader("📈 每日收益趋势")

            # 将日期转换为中文月日格式
            def to_chinese_date(d):
                try:
                    parts = str(d).strip().split('-')
                    if len(parts) >= 3:
                        return f"{int(parts[1])}月{int(parts[2])}日"
                    return str(d)
                except (ValueError, IndexError):
                    return str(d)

            chinese_dates = [to_chinese_date(d) for d in date_list]

            daily_revenues = [s['日收益_元'] for s in all_summaries]
            fig_daily_rev = go.Figure()
            fig_daily_rev.add_trace(go.Scatter(
                x=chinese_dates,
                y=daily_revenues,
                mode='lines+markers',
                name='日收益',
                line=dict(color='steelblue', width=2),
                marker=dict(size=4),
                fill='tozeroy',
                fillcolor='rgba(70,130,180,0.1)'
            ))
            # 添加均值参考线
            avg_daily = np.mean(daily_revenues)
            fig_daily_rev.add_hline(
                y=avg_daily, line_dash="dash", line_color="red",
                annotation_text=f"日均: {avg_daily:.0f} 元"
            )
            fig_daily_rev.update_layout(
                title="每日收益折线图",
                xaxis_title='日期',
                yaxis_title='收益 (元)',
                template='plotly_white',
                height=400,
                xaxis=dict(tickangle=45)
            )
            st.plotly_chart(fig_daily_rev, use_container_width=True)

            # 选择日期查看策略
            st.divider()
            st.subheader("📅 查看单日配储策略")

            date_options = [s['日期'] for s in all_summaries]

            # 用 session_state 记住选中的日期索引
            if "selected_date_idx" not in st.session_state or st.session_state.get("prev_num_days") != num_days:
                st.session_state.selected_date_idx = 0

            selected_date_idx = st.selectbox(
                "选择日期",
                options=range(len(date_options)),
                format_func=lambda x: date_options[x],
                index=st.session_state.selected_date_idx,
                key="date_selector"
            )
            st.session_state.selected_date_idx = selected_date_idx
            st.session_state.prev_num_days = num_days

            selected_date = date_options[selected_date_idx]
            selected_summary = all_summaries[selected_date_idx]

            # 显示该日统计
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("日收益", f"{selected_summary['日收益_元']:.2f} 元")
            with col2:
                st.metric("充电量", f"{selected_summary['充电量_kWh']:.2f} kWh")
            with col3:
                st.metric("放电量", f"{selected_summary['放电量_kWh']:.2f} kWh")

            # 显示该日详细数据
            day_results = [r for r in all_results if r['日期'] == selected_date]
            day_df = pd.DataFrame(day_results)

            # 充放电功率曲线（合并到一张图）
            fig_power = go.Figure()
            fig_power.add_trace(go.Scatter(
                x=day_df['时间'],
                y=day_df['充电功率_kW'],
                mode='lines',
                name='充电功率',
                line=dict(color='blue', width=2),
                fill='tozeroy',
                fillcolor='rgba(0,0,255,0.1)'
            ))
            fig_power.add_trace(go.Scatter(
                x=day_df['时间'],
                y=day_df['放电功率_kW'],
                mode='lines',
                name='放电功率',
                line=dict(color='red', width=2),
                fill='tozeroy',
                fillcolor='rgba(255,0,0,0.1)'
            ))
            fig_power.update_layout(
                title=f"{selected_date} 充放电功率曲线",
                xaxis_title='时间',
                yaxis_title='功率 (kW)',
                template='plotly_white',
                height=400,
                xaxis=dict(tickangle=45, tickvals=day_df['时间'][::8])
            )
            st.plotly_chart(fig_power, use_container_width=True)

            # 电池电量曲线
            fig_soc = go.Figure()
            fig_soc.add_trace(go.Scatter(
                x=day_df['时间'],
                y=day_df['电池电量_kWh'],
                mode='lines',
                name='电池电量',
                line=dict(color='green', width=2),
                fill='tozeroy',
                fillcolor='rgba(0,255,0,0.1)'
            ))
            fig_soc.update_layout(
                title=f"{selected_date} 电池电量曲线",
                xaxis_title='时间',
                yaxis_title='电池电量 (kWh)',
                template='plotly_white',
                height=350,
                xaxis=dict(tickangle=45, tickvals=day_df['时间'][::8])
            )
            st.plotly_chart(fig_soc, use_container_width=True)

            # 显示详细表格
            st.subheader("📊 详细数据表")
            st.dataframe(day_df, use_container_width=True, height=400)

            # 导出结果
            st.divider()
            st.subheader("💾 导出优化结果")

            results_df = pd.DataFrame(all_results)
            summary_df = pd.DataFrame(all_summaries)

            csv_data = results_df.to_csv(index=False, encoding='utf-8-sig')
            st.download_button(
                label="📥 下载详细结果 (CSV)",
                data=csv_data,
                file_name=f"{selected_station}_储能优化详细结果.csv",
                mime="text/csv"
            )

            summary_csv = summary_df.to_csv(index=False, encoding='utf-8-sig')
            st.download_button(
                label="📥 下载每日收益汇总 (CSV)",
                data=summary_csv,
                file_name=f"{selected_station}_每日收益汇总.csv",
                mime="text/csv"
            )

        elif st.session_state.opt_results is not None and params_changed:
            st.info("参数已变更，请点击上方「开始优化」按钮重新计算。")

if __name__ == "__main__":
    main()
